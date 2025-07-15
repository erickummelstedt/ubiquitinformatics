import json 
import logging
import copy
import sys
import ast
import numpy as np
from pathlib import Path
import pandas as pd
import ast 


import matplotlib.pyplot as plt
from matplotlib import font_manager

# Dynamically get the backend path relative to this file
current_file = Path(__file__).resolve()
project_root = current_file.parents[2]  # Go up to project root
sys.path.insert(0, str(project_root))
local_path = project_root / 'back_end'
sys.path.insert(0, str(local_path))


def plot_96wells(figure=1, figure_name = 'Test',colorbar_type= 'PuRd', cdata=None, sdata=None, bdata=None, bcolors=None, bmeans=None, **kwargs):
    # from https://github.com/jaumebonet/RosettaSilentToolbox/blob/master/rstoolbox/plot/experimental.py
    """Plot data of a 96 well plate into an equivalent-shaped plot.

    Allows up to three data sources at the same time comming from experiments performed in a
    `96 wells microplate <https://www.wikiwand.com/en/Microtiter_plate>`_. Two of this data
    sources can have to be numerical, and are represented as color and size, while an extra
    boolean dataset can be represented by the color of the border of each well.

    Plot is based on :func:`~matplotlib.pyplot.scatter`; some graphical properties to control
    the visuals (such as ``cmap``), can be provided through this function.

    :param cdata: Data contentainer to be represented by color coding. Has to contain 8 rows
        and 12 columns, like the 96 well plate. Contains **continuos numerical data**.
    :type cdata: :class:`~pandas.DataFrame`
    :param sdata: Data contentainer to be represented by size coding. Has to contain 8 rows
        and 12 columns, like the 96 well plate. Contains **continuos numerical data**.
    :type sdata: :class:`~pandas.DataFrame`
    :param bdata: Data contentainer to be represented by the edge color. Has to contain 8 rows
        and 12 columns, like the 96 well plate. Contains **boolean data**.
    :type bdata: :class:`~pandas.DataFrame`
    :param bcolors: List of the two colors to identify the differences in the border for binary
        data. It has to be a list of 2 colors only. First color represents :data:`True` instances
        while the second color is :data:`False`. Default is ``black`` and ``green``.
    :type bcolors: :func:`list` of :class:`str`
    :param bmeans: List with the meanings of the boolean condition (for the legend). First color
        represents :data:`True` instances while the second color is :data:`False`. Default is
        ``True`` and ``False``.
    :type bmeans: :func:`list` of :class:`str`

    :return: Union[:class:`~matplotlib.figure.Figure`, :class:`~matplotlib.axes.Axes`]

    :raises:
        :ValueError: If input data is not a :class:`~pandas.DataFrame`.
        :ValueError: If :class:`~pandas.DataFrame` do not has the proper shape.
        :ValueError: If ``bcolors`` of ``bmeans`` are provided with sizes different than 2.

    .. rubric:: Example

    .. ipython::
        :okwarning:

        In [1]: from rstoolbox.plot import plot_96wells
           ...: import numpy as np
           ...: import pandas as pd
           ...: import matplotlib.pyplot as plt
           ...: np.random.seed(0)
           ...: df = pd.DataFrame(np.random.randn(8, 12))
           ...: fig, ax = plot_96wells(cdata = df, sdata = -df, bdata = df<0)
           ...: plt.subplots_adjust(left=0.1, right=0.8, top=0.9, bottom=0.1)

        @savefig plot_96wells_docs.png width=5in
        In [2]: plt.show()

        In [3]: plt.close()
    """

    # Changes in one of this parameters should change others to ensure size fit.
    fig = plt.figure(figure, figsize=(15, 7))
    fig.suptitle(figure_name, fontsize=16)
    ax  = plt.subplot2grid((1, 1), (0, 0), fig=fig)
    top = 2000
    bot = 50
    sizeOfFont = 15
    ticks_font = font_manager.FontProperties(style='normal', size=sizeOfFont, weight='normal')

    # Fixed: THIS CANNOT CHANGE!
    kwargs['x'] = list(range(1, 13)) * 8
    kwargs['y'] = sorted(list(range(1, 9)) * 12)

    # Modified by the input data.
    kwargs.setdefault('s', [top, ] * len(kwargs['y']))
    kwargs.setdefault('c', 'white')
    kwargs.setdefault('edgecolor', ['black', ] * len(kwargs['y']))
    kwargs.setdefault('linewidths', 1.5)
    kwargs.setdefault('cmap', colorbar_type)

    # Color Data
    if cdata is not None:
        if not isinstance(cdata, pd.DataFrame):
            raise ValueError('Wrong data type')
        if cdata.shape != (8, 12):
            raise ValueError('Wrong data shape')
        kwargs['c'] = cdata.values.flatten()

    # Size Data
    if sdata is not None:
        if not isinstance(sdata, pd.DataFrame):
            raise ValueError('Wrong data type')
        if sdata.shape != (8, 12):
            raise ValueError('Wrong data shape')
        p = sdata.values.flatten()
        p = ((p - np.min(p)) / np.ptp(p)) * (top - bot) + bot
        kwargs['s'] = p

    # Border Data
    if bdata is not None:
        if not isinstance(bdata, pd.DataFrame):
            raise ValueError('Wrong data type')
        if bdata.shape != (8, 12):
            raise ValueError('Wrong data shape')
        if not (bdata.dtypes == bool).all():
            raise ValueError('bdata has to be booleans')
        if bcolors is None:
            bcolors = ['black', 'green']
        if bmeans is None:
            bmeans = ['True', 'False']
        if len(bcolors) < 2:
            raise ValueError('At least to border colors need to be provided')
        if len(bmeans) < 2:
            raise ValueError('At least to binary names need to be provided')
        b = bdata.values.flatten()
        b = [bcolors[0] if _ else bcolors[1] for _ in b]
        kwargs['edgecolor'] = b

    # PLOT
    mesh = ax.scatter(**kwargs)

    # Image aspects
    ax.grid(False)
    ax.set_xticks(range(1, 13))
    ax.xaxis.tick_top()
    ax.set_yticks(range(1, 9))
#    ax.set_yticklabels(string.ascii_uppercase[0:9])
    ax.set_ylim((8.5, 0.48))
    ax.set_aspect(1)
    ax.tick_params(axis='both', which='both', length=0)
    for spine in ax.spines.values():
        spine.set_visible(False)

    for label in ax.get_xticklabels():
        label.set_fontproperties(ticks_font)

    for label in ax.get_yticklabels():
        label.set_fontproperties(ticks_font)
    
    ## adding text to 96 well plate
    for text, x, y in zip(cdata.values.flatten(), list(range(1, 13)) * 8, sorted(list(range(1, 9)) * 12)):
        if text == 0: 
            new_text = ''
        else:
            new_text = text 
        if text < 10:
            plt.text(x-0.175, y+0.15, str(new_text), fontsize = 23)
        else:
            plt.text(x-0.275, y+0.15, str(new_text), fontsize = 23)

    return fig, ax


# ============================================
# Functions 
# ============================================

# This function creates an encoded dictionary that maps combinations of donors and enzymes to unique integers.
def create_e_d_encoded_dictionary(list_of_enzymes, list_of_donors):
    """
    Creates a dictionary that encodes the combinations of donors and enzymes.
    
    Parameters:
        list_of_enzymes (list): List of enzyme names.
        list_of_donors (list): List of donor names.
        
    Returns:
        dict: Dictionary mapping (donor, enzyme) tuples to unique integers.
    """
    e_d_encoded_dictionary = {}
    count = 1 
    for i in list_of_enzymes: 
        for j in list_of_donors:
            ## these are the self reactions should not be included
            if ((j == 'ubi_ubq_1_K48_SMAC') & (i=='Ubc13/Mms2')) | ((j == 'ubi_ubq_1_K63_SMAC') & (i=='Ube2K'))|((j == 'ubi_ubq_1_K63_SMAC') & (i=='gp78/Ube2g2')):
                count = count
            else:
                e_d_encoded_dictionary[j, i] = count 
                count = count+1 
    return e_d_encoded_dictionary  

def select_columns_by_keyword(combined_database, keyword):
    """
    Select columns from a DataFrame that contain a specific keyword (case-insensitive).

    Parameters:
        combined_database (pd.DataFrame): The input DataFrame.
        keyword (str): The keyword to search for in column names.

    Returns:
        list: A list of matching column names.
    """
    return [col for col in combined_database.columns if keyword.lower() in col.lower()]


def get_combined_data(combined_database, index_list, table_origin, keyword='formation'):
    """
    Returns formation-related columns for specified indices and table_origin,
    preserving the order of index_list.
    
    Parameters:
        combined_database (pd.DataFrame): The full database to filter from.
        index_list (list of int): List of index values to retrieve.
        table_origin (str): 'Reactions', 'Donors', etc.
        
    Returns:
        pd.DataFrame: Filtered DataFrame with only formation-related columns.
    """
    # Filter by condition
    filtered_rows = combined_database[
        (combined_database['index'].isin(index_list)) &
        (combined_database['table_origin'] == table_origin)
    ]

    # Preserve order of index_list
    filtered_rows_sorted = filtered_rows.set_index('index').loc[index_list].reset_index()

    # Select columns that contain 'formation' in the name
    formation_columns = select_columns_by_keyword(combined_database, keyword)
    
    # Extract only those columns
    return filtered_rows_sorted[formation_columns]

# Modified folding function: always creates two columns, padding with None if necessary
def fold_series(series, chunk_length=8, min_chunks=2):
    total_length = len(series)
    num_chunks = max(int(np.ceil(total_length / chunk_length)), min_chunks)
    padded = series.tolist() + [None] * (num_chunks * chunk_length - total_length)
    folded = [padded[i * chunk_length:(i + 1) * chunk_length] for i in range(num_chunks)]
    return pd.DataFrame(folded).transpose()

# Fold DataFrame function: applies folding to each column and renames columns for clarity
def fold_dataframe(df):
    """
    Apply folding to each column in a DataFrame using a provided fold_series function,
    then rename columns for clarity.

    Parameters:
        df (pd.DataFrame): The input DataFrame.
        fold_series (function): Function to apply to each column (e.g., for encoding).

    Returns:
        pd.DataFrame: The folded and renamed DataFrame.
    """
    # Check for maximum length constraint
    MAX_LENGTH = 16
    if len(df) > MAX_LENGTH:
        raise ValueError(f"DataFrame length {len(df)} exceeds maximum allowed length of {MAX_LENGTH}.")

    # Apply folding to each column
    folded_columns = [fold_series(df[col]) for col in df.columns]

    # Concatenate folded columns
    enzymes_donors_96 = pd.concat(folded_columns, axis=1, keys=df.columns)

    # Rename columns for clarity
    enzymes_donors_96.columns = [f"{col}_{i+1}" for col, i in enzymes_donors_96.columns]

    return enzymes_donors_96


def pad_and_validate_dataframe(input_df, required_columns=12):
    """
    Pad the DataFrame with empty columns if it has fewer than required_columns.
    Raise an error if it exceeds the limit. Convert NaNs to 0 and cast to integers.

    Parameters:
        input_df (pd.DataFrame): The input DataFrame to pad and validate.
        required_columns (int): The required number of columns (default: 12).

    Returns:
        pd.DataFrame: A cleaned and padded DataFrame with exactly required_columns columns.
    """
    current_columns = input_df.shape[1]

    # If fewer than required, add empty (None) columns
    if current_columns < required_columns:
        for i in range(current_columns + 1, required_columns + 1):
            input_df[f"padding_{i}"] = None

    # If more than required, raise an error
    elif current_columns > required_columns:
        raise ValueError(f"Final DataFrame has {current_columns} columns, which exceeds the {required_columns}-column limit.")

    # Convert the DataFrame to integers, filling NaNs with 0
    input_df = input_df.fillna(0).infer_objects().astype(int)

    return input_df

def inner_create_plate_dfs(data_dict, indexed_values, multimer_size=4):
    """    
    Creates a DataFrame with formation data for dimer, trimer, and tetramer levels,
    including encoded values based on donor and reaction combinations.  
    Parameters:
        combined_database (pd.DataFrame): The full database to filter from.
        indexed_values (list of int): List of index values for tetramers.
        e_d_encoded_dictionary (dict): Dictionary mapping (donor, reaction) tuples to
        unique integers.

    Returns:
        dict: A dictionary containing the encoded DataFrame and the encoded dictionary.
        The dictionary will have the following keys:
            - 'enzymes_donors_96': DataFrame with encoded donor and reaction formation
              data for dimer, trimer, and tetramer levels.
            - 'deprots_96': DataFrame with encoded deprotection data for dimer and trimer levels.
            - 'dimer_acceptors_96': DataFrame with encoded dimer acceptor data
              for dimer levels.
            - 'e_d_encoded_dictionary': Dictionary mapping (donor, reaction) tuples to
              unique integers for donor and reaction combinations.
            - 'deprot_encoded_dictionary': Dictionary mapping deprotection types to unique integers.
            - 'dimer_encoded_dictionary': Dictionary mapping dimer acceptor names to unique integers.
    """    
    # ================================
    # Extract the necessary data from the data_dict
    # ================================
    combined_database = data_dict['combined_database']
    context_history = data_dict['context_history']

    # ================================
    # Determine the levels of formation based on multimer size
    # ================================
    if multimer_size == 4:
        levels = ['dimer', 'trimer', 'tetramer']
    elif multimer_size == 5:
        levels = ['dimer', 'trimer', 'tetramer', 'pentamer']

    # ================================
    # Reactants and Donors
    # ================================
    # Inputs:
    # combined_database: The full database to filter from.
    # indexed_values: List of index values for tetramers.
    # e_d_encoded_dictionary: Dictionary mapping (donor, reaction) tuples to unique integers.
    # ================================

    # List of donors and enzymes to encode
    list_of_donors = [
        'ubi_ubq_1_K48_SMAC', 
        'ubi_ubq_1_K63_SMAC', 
        'ubi_ubq_1_K48_SMAC_K63_ABOC', 
        'ubi_ubq_1_K48_ABOC_K63_SMAC', 
        'ubi_ubq_1_K48_ABOC_K63_ABOC'
        ]

    list_of_enzymes = [
        'gp78/Ube2g2', 
        'Ube2K', 
        'Ubc13/Mms2'
        ]
    
    # Create the encoded dictionary
    e_d_encoded_dictionary = create_e_d_encoded_dictionary(list_of_enzymes, list_of_donors)

    # ================================
     # Create an empty DataFrame to hold the results
    enzymes_donors_96 = pd.DataFrame()

    # Filter the combined database for reactions and donors based on indexed values
    filtered_reactions = get_combined_data(combined_database, indexed_values, 'Reactions', 'formation')
    filtered_donors = get_combined_data(combined_database, indexed_values, 'Donors', 'formation')

    # Iterate through each level of formation
    for level in levels:
        if level == 'dimer':
            continue  # Skip this iteration
        else:
            # Create columns for donor and reaction formation data
            enzymes_donors_96[f'{level}_donor'] = filtered_donors[f'{level}_formation']
            enzymes_donors_96[f'{level}_reaction'] = filtered_reactions[f'{level}_formation']

            # Encode the donor and reaction combinations using the encoded dictionary
            # e_d_encoded mean enzyme donor encoded
            enzymes_donors_96[f'{level}_e_d_encoded'] = enzymes_donors_96.apply(
                lambda row: e_d_encoded_dictionary.get((row[f'{level}_donor'], row[f'{level}_reaction']), None), axis=1
            )

    # Pull only the encoded columns
    e_d_encoded_columns = select_columns_by_keyword(enzymes_donors_96, '_e_d_encoded')

    # Filter the DataFrame to keep only the encoded columns
    enzymes_donors_96 = enzymes_donors_96[e_d_encoded_columns]

    # Fold the DataFrame to ensure each column has a maximum of 8 values
    # Final DataFrame should have 12 columns, padded if necessary with encoded values
    enzymes_donors_96 = fold_dataframe(enzymes_donors_96)
    
    # ================================
    # Functions for Deprotections
    # ================================
    # Inputs:
    # filtered_reactions: The filtered reactions DataFrame containing deprotection information.
    # indexed_values: List of index values for tetramers.
    # levels: List of levels for deprotection (e.g., dimer, trimer, tetramer).
    # ================================
    
    # Create the encoded dictionary
    deprot_encoded_dictionary = {
        'SMAC_deprot': 1,
        'FAKE_deprot': 2
        }
    
    # Create an empty DataFrame to hold the results
    deprots_96 = pd.DataFrame()

    filtered_deprots = get_combined_data(combined_database, indexed_values, 'Reactions', 'deprotection')

    for level in levels[:-1]:  # Exclude 'final' level as the reactions end with a formation
        # Create columns for donor and reaction formation data
        deprots_96[f'{level}_deprotection'] = filtered_deprots[f'{level}_deprotection']

        # Encode the donor and reaction combinations using the encoded dictionary
        # e_d_encoded mean enzyme donor encoded
        deprots_96[f'{level}_deprot_encoded'] = deprots_96.apply(
            lambda row: deprot_encoded_dictionary.get((row[f'{level}_deprotection']), None), axis=1
        )
    
    # Pull only the encoded columns
    deprot_encoded_columns = select_columns_by_keyword(deprots_96, '_deprot_encoded')

    # Filter the DataFrame to keep only the encoded columns
    deprots_96 = deprots_96[deprot_encoded_columns]

    # Fold the DataFrame to ensure each column has a maximum of 8 values
    # Final DataFrame should have 12 columns, padded if necessary with encoded values
    deprots_96 = fold_dataframe(deprots_96)

    # ================================
    # Functions for Acceptors
    # ================================
    # Inputs:
    # context_history: The context history DataFrame containing acceptor information.   
    # indexed_values: List of index values for tetramers.
    # ================================

    # Create a dimer encoded dictionary
    dimer_encoded_dictionary = {
        'his-GG-1ubq-1-(<K48_1ubq-2-(<K48_SMAC><K63_ABOC>)>)' : 9,
        'his-GG-1ubq-1-(<K48_1ubq-2-(<K48_ABOC><K63_SMAC>)>)' : 10,
        'his-GG-1ubq-1-(<K48_1ubq-2-(<K48_ABOC><K63_ABOC>)>)' : 11,

        'his-GG-1ubq-1-(<K48_1ubq-2-(<K48_SMAC>)><K63_ABOC>)': 12,
        'his-GG-1ubq-1-(<K48_1ubq-2-(<K48_SMAC><K63_ABOC>)><K63_ABOC>)': 13,
        'his-GG-1ubq-1-(<K48_1ubq-2-(<K48_ABOC><K63_SMAC>)><K63_ABOC>)': 14,

        'his-GG-1ubq-1-(<K63_1ubq-2-(<K48_SMAC><K63_ABOC>)>)': 15,
        'his-GG-1ubq-1-(<K63_1ubq-2-(<K48_ABOC><K63_SMAC>)>)': 16,
        'his-GG-1ubq-1-(<K63_1ubq-2-(<K48_ABOC><K63_ABOC>)>)': 17,

        'his-GG-1ubq-1-(<K48_ABOC><K63_1ubq-2-(<K63_SMAC>)>)': 18,
        'his-GG-1ubq-1-(<K48_ABOC><K63_1ubq-2-(<K48_SMAC><K63_ABOC>)>)': 19,
        'his-GG-1ubq-1-(<K48_ABOC><K63_1ubq-2-(<K48_ABOC><K63_SMAC>)>)': 20
    }

    # Filter the combined database for reactions and donors based on indexed values
    # gonna have to be pulled from context_history
    filtered_acceptors = context_history[(context_history['index'].isin(indexed_values))]

    # Preserve order of index_list
    filtered_acceptors_sorted = filtered_acceptors.set_index('index').loc[indexed_values].reset_index()

    # Select columns that contain 'formation' in the name
    dimer_acceptor_columns = select_columns_by_keyword(context_history, 'dimer_formation')
    
    # Get the dimer acceptor formation data
    dimer_acceptors_96 = filtered_acceptors_sorted[dimer_acceptor_columns]

    # Get the multimer string name 
    dimer_acceptors_96['dimer_names'] = dimer_acceptors_96['dimer_formation'].apply(
        lambda x: ast.literal_eval(x).get('multimer_string_name') if isinstance(x, str) else None
        )

    # Create acceptor encoded plate maps
    dimer_acceptors_96['dimers_encoded'] = dimer_acceptors_96['dimer_names'].map(dimer_encoded_dictionary)

    # Filter the DataFrame to keep only the encoded columns
    dimer_acceptors_96 = dimer_acceptors_96[['dimers_encoded']]

    # Fold the DataFrame to ensure each column has a maximum of 8 values
    # Final DataFrame should have 12 columns, padded if necessary with encoded values
    dimer_acceptors_96 = fold_dataframe(dimer_acceptors_96)
   
    # ================================
    # Functions for Padding and Finalization
    # ================================

    output_enzymes_donors_96 = pad_and_validate_dataframe(enzymes_donors_96, required_columns=12)
    output_deprots_96 = pad_and_validate_dataframe(deprots_96, required_columns=12)
    output_dimer_acceptors_96 = pad_and_validate_dataframe(dimer_acceptors_96, required_columns=12)
    
    # ================================
    # Prepare the output dictionary
    # This dictionary will contain the encoded DataFrame and the encoded dictionary
    # It can be used for further processing or analysis.
    # The dimer_acceptors can be used for acceptor plate maps or other analyses.
    # ================================
    output_dict = {
        'enzymes_donors_96': output_enzymes_donors_96,
        'deprots_96': output_deprots_96,
        'dimer_acceptors_96': output_dimer_acceptors_96,
        'e_d_encoded_dictionary': e_d_encoded_dictionary,
        'deprot_encoded_dictionary': deprot_encoded_dictionary,
        'dimer_encoded_dictionary': dimer_encoded_dictionary
    }

    return output_dict

# ============================================
# Function to create an Excel file with the plate dataframes
# ============================================

# create_xlsx_bytes function that 
def create_xlsx_bytes(output_dict):

    """ 
    
    """
    from collections import defaultdict

    # Local variables for create_xlsx_bytes
    # List of components to count in the mixtures
    components = [
        'Ubc13/Mms2',
        'gp78/Ube2g2',
        'Ube2K',
        'ubi_ubq_1_K48_ABOC_K63_ABOC',
        'ubi_ubq_1_K48_SMAC_K63_ABOC',
        'ubi_ubq_1_K48_ABOC_K63_SMAC',
        'ubi_ubq_1_K48_SMAC',
        'ubi_ubq_1_K63_SMAC'
    ]
    default_final_concentrations = {
        # E1 enzyme
        'hUba1': 1,  # in µM
        # E2 enzymes
        'Ubc13/Mms2': 20,  # in µM
        'gp78/Ube2g2': 20,  # in µM
        'Ube2K': 20,  # in µM
        'ubi_ubq_1_K48_ABOC_K63_ABOC': 20, # in µM
        'ubi_ubq_1_K48_SMAC_K63_ABOC': 20, # in µM
        'ubi_ubq_1_K48_ABOC_K63_SMAC': 20, # in µM
        'ubi_ubq_1_K48_SMAC': 20, # in µM
        'ubi_ubq_1_K63_SMAC': 20, # in µM
        # Acceptors
        'Acceptor Dimer Concentration': 10,  # in nmol
        'Reaction Volume [uL]': 200,  # in µL
        # TCEP & ATP
        'TCEP': 1,  # in mM 
        'ATP/Mg2+': 10,  # in mM
    }

    initial_concentraions = {
        'TCEP': 10,  # in mM 
        'ATP/Mg2+': 100,  # in mM
        'hUba1': 10,  # in µM
    }

    masses = {
        'hUba1': 116300,  # in Da
        'Ubc13/Mms2': 33800,  # in Da
        'gp78/Ube2g2': 27600,  # in Da
        'Ube2K': 22400,  # in Da
        # Acceptor monomers
        'histag_ubi_ubq_1' : 9527, # in Da
        'histag_ubi_ubq_1_K48_aboc' : 9658, # in Da
        'histag_ubi_ubq_1_K63_aboc' : 9658, # in Da
        # Donors 
        'ubi_ubq_1_K48_ABOC_K63_ABOC': 8851, # in Da 
        'ubi_ubq_1_K48_SMAC_K63_ABOC': 8864, # in Da
        'ubi_ubq_1_K48_ABOC_K63_SMAC': 8864, # in Da
        'ubi_ubq_1_K48_SMAC': 8733.00, # in Da
        'ubi_ubq_1_K63_SMAC': 8733.00, # in Da
    }

    # Define acceptor description dictionary
    acceptor_description_dict = [
        {"No.": "Ub₂ᴬ 9", "Dimer Linkage": "K48", "Proximal Ub K48": "-", "Proximal Ub K63": "K", "Distal Ub K48": "Smac", "Distal Ub K63": "Aboc", "MW (Da)": 18372},
        {"No.": "Ub₂ᴬ 10", "Dimer Linkage": "K48", "Proximal Ub K48": "-", "Proximal Ub K63": "K", "Distal Ub K48": "Aboc", "Distal Ub K63": "Smac", "MW (Da)": 18372},
        {"No.": "Ub₂ᴬ 11", "Dimer Linkage": "K48", "Proximal Ub K48": "-", "Proximal Ub K63": "K", "Distal Ub K48": "Aboc", "Distal Ub K63": "-", "MW (Da)": 18360},
        {"No.": "Ub₂ᴬ 12", "Dimer Linkage": "K48", "Proximal Ub K48": "-", "Proximal Ub K63": "Aboc", "Distal Ub K48": "Smac", "Distal Ub K63": "K", "MW (Da)": 18372},
        {"No.": "Ub₂ᴬ 13", "Dimer Linkage": "K48", "Proximal Ub K48": "-", "Proximal Ub K63": "Aboc", "Distal Ub K48": "Smac", "Distal Ub K63": "-", "MW (Da)": 18504},
        {"No.": "Ub₂ᴬ 14", "Dimer Linkage": "K48", "Proximal Ub K48": "-", "Proximal Ub K63": "Aboc", "Distal Ub K48": "Aboc", "Distal Ub K63": "-", "MW (Da)": 18504},
        {"No.": "Ub₂ᴬ 15", "Dimer Linkage": "K63", "Proximal Ub K48": "K", "Proximal Ub K63": "-", "Distal Ub K48": "Smac", "Distal Ub K63": "Aboc", "MW (Da)": 18372},
        {"No.": "Ub₂ᴬ 16", "Dimer Linkage": "K63", "Proximal Ub K48": "K", "Proximal Ub K63": "-", "Distal Ub K48": "Aboc", "Distal Ub K63": "Smac", "MW (Da)": 18372},
        {"No.": "Ub₂ᴬ 17", "Dimer Linkage": "K63", "Proximal Ub K48": "K", "Proximal Ub K63": "-", "Distal Ub K48": "Aboc", "Distal Ub K63": "-", "MW (Da)": 18360},
        {"No.": "Ub₂ᴬ 18", "Dimer Linkage": "K63", "Proximal Ub K48": "Aboc", "Proximal Ub K63": "-", "Distal Ub K48": "K", "Distal Ub K63": "-", "MW (Da)": 18372},
        {"No.": "Ub₂ᴬ 19", "Dimer Linkage": "K63", "Proximal Ub K48": "Aboc", "Proximal Ub K63": "-", "Distal Ub K48": "Smac", "Distal Ub K63": "-", "MW (Da)": 18504},
        {"No.": "Ub₂ᴬ 20", "Dimer Linkage": "K63", "Proximal Ub K48": "Aboc", "Proximal Ub K63": "-", "Distal Ub K48": "Aboc", "Distal Ub K63": "Smac", "MW (Da)": 18504},
    ]

    # prep_headers change depending on the regent
    donor_id_dict = {
        'Ubc13/Mms2' : 'Enzyme',
        'gp78/Ube2g2' : 'Enzyme',
        'Ube2K' : 'Enzyme',
        'ubi_ubq_1_K48_ABOC_K63_ABOC' : 'Donor: Ubᴰ 5',
        'ubi_ubq_1_K48_SMAC_K63_ABOC' : 'Donor: Ubᴰ 2',
        'ubi_ubq_1_K48_ABOC_K63_SMAC' : 'Donor: Ubᴰ 4',
        'ubi_ubq_1_K48_SMAC' : 'Donor: Ubᴰ 1',
        'ubi_ubq_1_K63_SMAC' : 'Donor: Ubᴰ 3',
    }

    # ============================================================
    # Extra functions
    # ============================================================

    # Create plate dataframes for the selected indexes
    def count_nonzero_values(df):
        """
        Count occurrences of non-zero values in a DataFrame.

        Parameters:
            df (pd.DataFrame): The input DataFrame.

        Returns:
            dict: Dictionary of value counts, excluding zeros.
        """
        counts = pd.Series(df.values.ravel())
        return counts[counts != 0].value_counts().sort_index().to_dict()

    def count_enzyme_donor_and_components(e_d_encoded_dict, count_dict):

        """ Count the occurrences of each mixture and its components in the provided count dictionary.
        Args:
            e_d_encoded_dict (dict): Dictionary mapping encoded mixtures to their original tuples.
            count_dict (dict): Dictionary with encoded keys and their counts.
        Returns:
            tuple: A tuple containing:
                - enzyme_donor_counts (dict): Dictionary with counts of each mixture.
                - component_counts (dict): Dictionary with counts of each component.
                - enzyme_donor_code_map (dict): Mapping from mixture tuples to their encoded codes.
        """

        # List of components to count in the mixtures
        components = [
            'Ubc13/Mms2',
            'gp78/Ube2g2',
            'Ube2K',
            'ubi_ubq_1_K48_ABOC_K63_ABOC',
            'ubi_ubq_1_K48_SMAC_K63_ABOC',
            'ubi_ubq_1_K48_ABOC_K63_SMAC',
            'ubi_ubq_1_K48_SMAC',
            'ubi_ubq_1_K63_SMAC'
        ]

        # Count per mixture (decoded from keys)
        enzyme_donor_counts = defaultdict(int)
        component_counts = defaultdict(int)

        # Reverse lookup: map encoded values back to mixture keys
        reversed_dict = {v: k for k, v in e_d_encoded_dict.items()}

        # Ensure every mixture from the dictionary is present, even if not in count_dict
        for encoded_key, mixture in reversed_dict.items():
            enzyme_donor_counts[mixture] = 0  # Initialize all mixtures to 0

        for encoded_key, count in count_dict.items():
            if encoded_key not in reversed_dict:
                continue  # skip unknown keys
            mixture = reversed_dict[encoded_key]
            enzyme_donor_counts[mixture] += count
            # Flatten the components in the mixture (tuple of (substrate, enzyme))
            for item in mixture:
                for component in components:
                    if component == item or (isinstance(item, str) and component in item):
                        component_counts[component] += count

        # Also return the mapping from mixture (tuple) to encoded code
        enzyme_donor_code_map = {mixture: code for code, mixture in e_d_encoded_dict.items()}
        return dict(enzyme_donor_counts), dict(component_counts), enzyme_donor_code_map

    # ===========================================================
    # Set up functions logic
    # ===========================================================

    # Extract the plate dataframes from the output dictionary
    enzymes_donors_96 = output_dict['enzymes_donors_96']
    deprots_96 = output_dict['deprots_96']
    acceptors_96 = output_dict['dimer_acceptors_96']

    # Count non-zero values in each DataFrame
    enzymes_donors_96_counts = count_nonzero_values(enzymes_donors_96)
    deprots_96_counts = count_nonzero_values(deprots_96)
    acceptors_96_counts = count_nonzero_values(acceptors_96)

    # pull dictionaries from the output_dict
    e_d_encoded_dictionary = output_dict['e_d_encoded_dictionary']
    deprot_encoded_dictionary = output_dict['deprot_encoded_dictionary']
    dimer_encoded_dictionary = output_dict['dimer_encoded_dictionary']

    enzyme_donor_counts, component_counts, enzyme_donor_code_map = count_enzyme_donor_and_components(
        e_d_encoded_dictionary, enzymes_donors_96_counts)

    # ===========================================================
    # Inner function to create the Excel file bytes
    # ===========================================================

    def inner_create_xlsx_bytes(enzyme_donor_counts, component_counts):
        
        """
        Create an Excel file with two sheets: one for enzyme_donor_counts and one for component_counts.
        Ensures all components from components are present in the component sheet (with 0 if missing).
        Includes mixture codes in the mixture sheet if provided.
        Adds a sum row at the bottom of the mixture counts.
        """
        
        # Mixed Reagents DataFrame
        enzyme_donor_data = []
        for index, (mixture, count) in enumerate(enzyme_donor_counts.items()):
            code = e_d_encoded_dictionary.get(mixture)  # Get the value from the encoding dictionary
            reaction_volume = default_final_concentrations.get('Reaction Volume [uL]', 200)
            enzyme_donor_data.append({'Donor + Enzyme Mix': mixture, 'Donor + Enzyme Id': code, 'Count': count, 'Reaction Volume (uL)': reaction_volume})
        enzyme_donor_df = pd.DataFrame(enzyme_donor_data, columns=['Donor + Enzyme Mix', 'Donor + Enzyme Id', 'Count', 'Reaction Volume (uL)'])
        # Add sum row
        sum_row = {'Donor + Enzyme Mix': 'TOTAL', 'Donor + Enzyme Id': '', 'Count': enzyme_donor_df['Count'].sum(), 'Reaction Volume (uL)': ''}
        enzyme_donor_df = pd.concat([enzyme_donor_df, pd.DataFrame([sum_row])], ignore_index=True)
        # Add Excel formula for 'Donor + Enzyme Volume (uL)' column
        from openpyxl import load_workbook
        
        # Start column index for the mixture sheet
        start_column_index = 1  # Start from the second column (B) (0-indexed, so 1 is column B)
        start_row_index = 1  # Start from the second row (0-indexed, so 1 is row 2)
        start_row = start_row_index + 1

        from io import BytesIO
        output = BytesIO()
        # Create the Excel file with pandas
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            enzyme_donor_df.to_excel(writer, sheet_name='Mixed Reagents Cal.', index=False, startcol=start_column_index, startrow=start_row_index)
            # Write the component sheet as before
            if components is not None:
                comp_data = [(comp, component_counts.get(comp, 0)) for comp in components]
                component_df = pd.DataFrame(comp_data, columns=['Base Reagent', 'Count'])
            else:
                component_df = pd.DataFrame(list(component_counts.items()), columns=['Base Reagent', 'Count'])
            component_df.to_excel(writer, sheet_name='Base Reagent Cal.', index=False, startcol=start_column_index, startrow=start_row_index)

        # Now load the workbook from the BytesIO buffer
        output.seek(0)
        wb = load_workbook(output)
        ws = wb['Mixed Reagents Cal.']

        # Insert new column for 'Donor + Enzyme Volume (uL)\n(Count * Reaction Volume (uL) * 1.1)' after 'Reaction Volume (uL)'
        insert_at = len(enzyme_donor_df.columns) + start_column_index + 1 # After 4th column (A-D: Mix, Id, Count, Reaction Vol)
        ws.insert_cols(insert_at)
        ws.cell(row=start_row, column=insert_at).value = 'Donor + Enzyme Volume (uL)\n(Count * Reaction Volume (uL) * 1.1)'
        
        # Get 'Reaction Volume (uL)' and 'Count' column indexes
        reaction_vol_col_idx = start_column_index + 4  # Assuming 'Reaction Volume (uL)' is the 4th column (D)   
        count_col_idx = start_column_index + 3  # Assuming 'Count' is the 3rd column (C) 

        # Add formula for each row except the sum row
        for row in range(start_row + 1, ws.max_row):
            count_cell = ws.cell(row=row, column=reaction_vol_col_idx).coordinate  # 'Count' column
            reaction_vol_cell = ws.cell(row=row, column=count_col_idx).coordinate  # 'Reaction Volume (uL)' column
            formula = f"={count_cell}*{reaction_vol_cell}*1.1"
            ws.cell(row=row, column=insert_at).value = formula
        # --- Add new columns for Donor, Donor Volume (uL), Enzyme, Enzyme Volume (uL) as Excel formulas ---
        # Insert Donor and Enzyme columns after 'Donor + Enzyme Volume (uL)\n(Count * Reaction Volume (uL) * 1.1)'
        donor_col = insert_at + 1
        donor_id_col = donor_col + 1
        donor_vol_col = donor_id_col + 1
        enzyme_col = donor_vol_col + 1
        enzyme_vol_col = enzyme_col + 1

        # Insert columns for Donor, Donor Id, Donor Volume, Enzyme, Enzyme Volume
        ws.insert_cols(donor_col, amount=enzyme_vol_col-insert_at)  # Insert 5 columns for Donor, Donor Id, Donor Volume, Enzyme, Enzyme Volume
        ws.cell(row=start_row, column=donor_col).value = 'Donor'    
        ws.cell(row=start_row, column=donor_id_col).value = 'Donor Id'
        ws.cell(row=start_row, column=donor_vol_col).value = 'Donor Volume (uL)'
        ws.cell(row=start_row, column=enzyme_col).value = 'Enzyme'
        ws.cell(row=start_row, column=enzyme_vol_col).value = 'Enzyme Volume (uL)'

        # Fill Donor and Enzyme columns by parsing the tuple string
        import ast
        for row in range(start_row + 1, ws.max_row):
            mix_val = ws.cell(row=row, column=start_column_index+1).value
            try:
                t = ast.literal_eval(mix_val) if isinstance(mix_val, str) and mix_val.startswith('(') else ('','')
                donor, enzyme = t if isinstance(t, tuple) and len(t) == 2 else ('','')
            except Exception:
                donor, enzyme = '', ''
            ws.cell(row=row, column=donor_col).value = donor
            # Set the donor ID based on the donor name
            ws.cell(row=row, column=donor_id_col).value = donor_id_dict.get(donor)
            ws.cell(row=row, column=enzyme_col).value = enzyme
            # Set Excel formulas for volumes: (Count * Reaction Volume (uL) * 1.1) / 2
            count_cell = ws.cell(row=row, column=count_col_idx).coordinate
            reaction_vol_cell = ws.cell(row=row, column=reaction_vol_col_idx).coordinate
            # Set formulas for Donor Volume and Enzyme Volume
            ws.cell(row=row, column=donor_vol_col).value = f"=({count_cell}*{reaction_vol_cell}*1.1)/2"
            ws.cell(row=row, column=enzyme_vol_col).value = f"=({count_cell}*{reaction_vol_cell}*1.1)/2"
        
        # Update header formatting and auto-fit for all columns (including new ones)
        from openpyxl.styles import Font, Border, Side, Alignment
        header_font = Font(bold=True)
        border = Border(bottom=Side(border_style="medium"))#, left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"))
        for col in range(start_column_index + 1, ws.max_column + 1):
            cell = ws.cell(row=start_row, column=col)
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        # Auto-fit all columns to the width of their titles or widest value, accounting for wrapping
        for index, col in enumerate(ws.columns):
            if index < start_column_index:
                continue # Skip the columns before the start_column_index
            header_cell = col[start_column_index]  # Header cell is the first row in the column
            col_letter = header_cell.column_letter
            if header_cell.value == 'Mass Required (mg)\n(MW (Da) * Amount Needed (nmol) / 1000)':
                ws.column_dimensions[col_letter].width = 10  # Set a tighter width for this column
                continue
            if '\n' in str(header_cell.value):
                lines = str(header_cell.value).split('\n')
                max_length = max(len(line) for line in lines)
            else:
                max_length = len(str(header_cell.value)) if header_cell.value is not None else 0
            for cell in col[start_column_index+1:]:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width
        # Highlight non-zero rows in specified columns with light green fill (excluding 'Total' row and 'Donor + Enzyme Mix')
        from openpyxl.styles import PatternFill
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        # Get column indices for the specified columns (excluding 'Donor + Enzyme Mix')
        col_names = [
            'Donor + Enzyme Id', 'Donor', 'Donor Id','Donor Volume (uL)', 'Enzyme', 'Enzyme Volume (uL)'
        ]
        col_indices = {}
        for col in range(start_column_index + 1, ws.max_column + 1):
            header = ws.cell(row=start_row, column=col).value
            if header in col_names:
                col_indices[header] = col
        # Highlight rows where 'Count' (col 3) is nonzero and not the sum row
        for row in range(start_row+1, ws.max_row+1):
            count_val = ws.cell(row=row, column=count_col_idx).value
            first_col_val = ws.cell(row=row, column=start_column_index+1).value
            try:
                is_nonzero = float(count_val) != 0
            except Exception:
                is_nonzero = False
            is_total = str(first_col_val).strip().upper() == 'TOTAL'
            if is_nonzero and not is_total:
                for colname in col_names:
                    col = col_indices.get(colname)
                    if col:
                        cell = ws.cell(row=row, column=col)
                        cell.fill = green_fill
                        cell.font = Font(bold=True)  # Make highlighted cell bold
        
        count_idx = 2
        donor_idx = 3
        reaction_vol_idx = donor_idx + 1
        total_vol_idx = reaction_vol_idx + 1
        final_conc_idx = total_vol_idx + 1
        amount_needed_idx = final_conc_idx + 1
        mw_idx = amount_needed_idx + 1
        mass_required_idx = mw_idx + 1

        # Format Base Reagent Cal. sheet
        ws_comp = wb['Base Reagent Cal.']

        # Insert 'Donor ID' column after Cal.
        ws_comp.insert_cols(start_column_index + donor_idx)  # Insert after 'Count' (which is now at col 2)
        ws_comp.cell(row=start_row_index + 1, column=start_column_index + donor_idx).value = 'Donor Id'
        # Insert 'Reaction Volume (uL)' column after 'Donor Id'
        ws_comp.insert_cols(start_column_index + reaction_vol_idx)  # Insert after 'Count' (which is now at col 2)
        ws_comp.cell(row=start_row_index + 1, column=start_column_index + reaction_vol_idx).value = 'Reaction Volume (uL)'
        # Insert 'Total Volume (uL)\n(Count * Reaction Volume (uL) * 1.1 / 2)' after 'Reaction Volume (uL)'
        ws_comp.insert_cols(start_column_index + total_vol_idx)
        ws_comp.cell(row=start_row_index + 1, column=start_column_index + total_vol_idx).value = 'Total Volume (uL)\n(Count * Reaction Volume (uL) * 1.1 / 2)'
        # Insert 'Final Concentration (uM)' after 'Total Volume (uL)...'
        ws_comp.insert_cols(start_column_index + final_conc_idx)
        ws_comp.cell(row=start_row_index + 1, column=start_column_index + final_conc_idx).value = 'Final Concentration (uM)'
        # Insert 'Amount Needed (nmol)\n(Count * Reaction Volume (uL)\n* Final Concentration (uM) * 1.1 / 1000)' after Final Concentration
        ws_comp.insert_cols(start_column_index + amount_needed_idx)
        ws_comp.cell(row=start_row_index + 1, column=start_column_index + amount_needed_idx).value = 'Amount Needed (nmol)\n(Count * Reaction Volume (uL)\n* Final Concentration (uM) * 1.1 / 1000)'
        # Insert 'MW (Da)' as the last column
        ws_comp.insert_cols(start_column_index + mw_idx)
        ws_comp.cell(row=start_row_index + 1, column=start_column_index + mw_idx).value = 'MW (Da)'
        # Insert 'Mass Required (ug)' as the next column
        ws_comp.insert_cols(start_column_index + mass_required_idx)
        ws_comp.cell(row=start_row_index + 1, column=start_column_index + mass_required_idx).value = 'Mass Required (mg)\n(MW (Da) * Amount Needed (nmol) / 1000)'
        # Set values for all rows except header
        for row in range(start_row+1, ws_comp.max_row+1):
            comp = ws_comp.cell(row=row, column=start_column_index + 1).value
            # Set the 'Donor' and 'Donor Id' based on the component
            reagent = ws_comp.cell(row=row, column=start_column_index + donor_idx - 2).value
            reagent_id = donor_id_dict.get(reagent)
            cleaned_id = reagent_id.replace("Enzyme", "").replace("Donor: ", "").strip()
            ws_comp.cell(row=row, column=start_column_index + donor_idx).value = cleaned_id
            # Set the 'Reaction Volume (uL)' based on the default final concentrations
            reaction_volume = default_final_concentrations.get('Reaction Volume [uL]', 200)
            ws_comp.cell(row=row, column=start_column_index + reaction_vol_idx).value = reaction_volume
            # Set the 'Count' value
            count_cell = ws_comp.cell(row=row, column=start_column_index + count_idx).coordinate  # 'Count' now at col 2
            reaction_vol_cell = ws_comp.cell(row=row, column=start_column_index + reaction_vol_idx).coordinate  # 'Reaction Volume (uL)' now at col 3
            # Set Total Volume (uL) as Excel formula: (Count * Reaction Volume (uL) * 1.1) / 2
            ws_comp.cell(row=row, column=start_column_index + total_vol_idx).value = f"=({count_cell}*{reaction_vol_cell}*1.1)/2"
            final_conc_cell = ws_comp.cell(row=row, column=start_column_index + final_conc_idx).coordinate
            # Set Final Concentration (uM) based on the default final concentrations
            ws_comp.cell(row=row, column=start_column_index + amount_needed_idx).value = f"=({count_cell}*{reaction_vol_cell}*{final_conc_cell}*1.1/1000)"
            ws_comp.cell(row=row, column=start_column_index + final_conc_idx).value = default_final_concentrations.get(comp, '')
            # Set mass value in the last column
            ws_comp.cell(row=row, column=start_column_index + mw_idx).value = masses.get(comp, '')
            # Set Mass Required (ug) as Excel formula: (MW (Da) * Amount Needed (nmol)) / 1,000,000, rounded to 3 decimals
            mass_cell = ws_comp.cell(row=row, column=start_column_index + mw_idx).coordinate
            amount_needed_cell = ws_comp.cell(row=row, column=start_column_index + amount_needed_idx).coordinate
            ws_comp.cell(row=row, column=start_column_index + mass_required_idx).value = f"=IF(AND(ISNUMBER({mass_cell}),ISNUMBER({amount_needed_cell})),ROUND({mass_cell}*{amount_needed_cell}/1000000,3),"")"
        # Header formatting: bold, centered, bordered, wrap text
        from openpyxl.styles import Font, Border, Side, Alignment
        header_font = Font(bold=True)
        border = Border(bottom=Side(border_style="medium")) # if you want more borders insert here, left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"))
        for col in range(start_column_index + 1, ws_comp.max_column + 1):
            cell = ws_comp.cell(row=start_row_index + 1, column=col)
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        # Auto-fit all columns to the width of their titles or widest value
        # Auto-fit all columns to the width of their titles or widest value, accounting for wrapping
        for index, col in enumerate(ws_comp.columns):
            if index < start_column_index:
                continue # Skip the columns before the start_column_index
            header_cell = col[start_column_index]
            col_letter = header_cell.column_letter
            # Special case for 'Mass Required (mg)\n(MW (Da) * Amount Needed (nmol) / 1000)' column
            if header_cell.value == 'Mass Required (mg)\n(MW (Da) * Amount Needed (nmol) / 1000)':
                ws_comp.column_dimensions[col_letter].width = 30  # Set a tighter width for this column
                continue
            if '\n' in str(header_cell.value):
                lines = str(header_cell.value).split('\n')
                max_length = max(len(line) for line in lines)
            else:
                max_length = len(str(header_cell.value)) if header_cell.value is not None else 0
            for cell in col[start_column_index + 1:]:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            adjusted_width = max_length + 2
            ws_comp.column_dimensions[col_letter].width = adjusted_width




            
        # --- Create new worksheet for base reagent preparation in the same workbook ---
        # Remove existing 'Base Reagent Prep' sheet if it exists
        if 'Base Reagent Prep' in wb.sheetnames:
            std = wb['Base Reagent Prep']
            wb.remove(std)
        ws_prep = wb.create_sheet('Base Reagent Prep')
        # Place initial concentrations for hUba1, TCEP, ATP/Mg2+ in the top left corner with units, starting at cell B2
        start_row, start_col = 2, 2  # B2
        ws_prep.cell(row=start_row, column=start_col).value = 'Stock Conc.'
        ws_prep.cell(row=start_row+1, column=start_col).value = 'hUba1 (uM)'
        ws_prep.cell(row=start_row+2, column=start_col).value = 'TCEP (mM)'
        ws_prep.cell(row=start_row+3, column=start_col).value = 'ATP/Mg2+ (mM)'
        ws_prep.cell(row=start_row+1, column=start_col+1).value = initial_concentraions.get('hUba1', '') if 'hUba1' in initial_concentraions else ''
        ws_prep.cell(row=start_row+2, column=start_col+1).value = initial_concentraions.get('TCEP', '')
        ws_prep.cell(row=start_row+3, column=start_col+1).value = initial_concentraions.get('ATP/Mg2+', '')
        # Expand columns and center the writing
        from openpyxl.styles import Alignment, Border, Side, Font
        for row in range(start_row, start_row+4):
            for col in range(start_col, start_col+2):
                cell = ws_prep.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True) if row == start_row else Font(bold=False)
        # Set column widths for better appearance
        ws_prep.column_dimensions['B'].width = 18
        ws_prep.column_dimensions['C'].width = 14
        # Add a medium border under the title row, but remove the right border from the value column
        for col in range(start_col, start_col+2):
            cell = ws_prep.cell(row=start_row, column=col)
            if col == start_col+1:
                cell.border = Border(
                    top=cell.border.top,
                    left=cell.border.left,
                    right=Side(border_style=None),
                    bottom=Side(border_style="medium")
                )
            else:
                cell.border = Border(
                    top=cell.border.top,
                    left=cell.border.left,
                    right=cell.border.right,
                    bottom=Side(border_style="medium")
                )
        # Remove the right border from all cells in value column
        for row in range(start_row, start_row+4):
            cell = ws_prep.cell(row=row, column=start_col+1)
            cell.border = Border(
                top=cell.border.top,
                left=cell.border.left,
                right=Side(border_style=None),
                bottom=cell.border.bottom
            )
        # Draw a medium border around the whole table (B2:C5)
        for row in range(start_row, start_row+4):
            for col in range(start_col, start_col+2):
                cell = ws_prep.cell(row=row, column=col)
                border_args = {}
                if row == start_row:
                    border_args['top'] = Side(border_style="medium")
                if row == start_row+3:
                    border_args['bottom'] = Side(border_style="medium")
                if col == start_col:
                    border_args['left'] = Side(border_style="medium")
                if col == start_col+1:
                    border_args['right'] = Side(border_style="medium")
                # Merge with any existing border
                cell.border = Border(
                    left=border_args.get('left', cell.border.left),
                    right=border_args.get('right', cell.border.right),
                    top=border_args.get('top', cell.border.top),
                    bottom=border_args.get('bottom', cell.border.bottom)
                )

        # Add a new table under Stock Conc. in B7 for hUba1 needs
        hUba1_needs_start_row = 7  # Row 7 (B7)
        hUba1_needs_start_col = 2  # Column B
        hUba1_needs_labels = [
            'Total hUba1 Needs',
            'Volume (uL)',
            'nmol',
            'MW (Da)',
            'Mass (mg)'
        ]
        for i, label in enumerate(hUba1_needs_labels):
            ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col).value = label
            # Only bold the first label ('Total hUba1 Needs'), others not bold
            ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col).font = Font(bold=(i == 0))
            # Center the writing in column B7 to B11 (label column)
            ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col).alignment = Alignment(horizontal='center', vertical='center')
            # For 'Volume (uL)' row, insert formula summing G14, G32, G50
            if label == 'Volume (uL)':
                ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col + 1).value = '=G14+G32+G50'
            # For 'nmol' row, insert formula: =C3*C8/1000
            elif label == 'nmol':
                ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col + 1).value = '=C3*C8/1000'
            # For 'MW (Da)' row, insert value from masses['hUba1']
            elif label == 'MW (Da)':
                ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col + 1).value = masses['hUba1']
            # For 'Mass (mg)' row, insert formula: =ROUND(C9*C10/1E6, 3)
            elif label == 'Mass (mg)':
                ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col + 1).value = '=ROUND(C9*C10/1E6, 3)'
            else:
                ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col + 1).value = ''
            ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col + 1).alignment = Alignment(horizontal='center', vertical='center')
        # Draw a border around the hUba1 needs box (B7:C10) and underline under the top row (B7:C7)
        from openpyxl.styles import Border, Side
        border_side = Side(border_style="medium")
        underline_side = Side(border_style="medium")
        for row in range(hUba1_needs_start_row, hUba1_needs_start_row + 5):
            for col in range(hUba1_needs_start_col, hUba1_needs_start_col + 2):
                cell = ws_prep.cell(row=row, column=col)
                border_args = {}
                if row == hUba1_needs_start_row:
                    border_args['top'] = border_side
                    border_args['bottom'] = underline_side  # underline under top row
                if row == hUba1_needs_start_row + 4:
                    border_args['bottom'] = border_side
                if col == hUba1_needs_start_col:
                    border_args['left'] = border_side
                if col == hUba1_needs_start_col + 1:
                    border_args['right'] = border_side
                # Merge with any existing border
                cell.border = Border(
                    left=border_args.get('left', cell.border.left),
                    right=border_args.get('right', cell.border.right),
                    top=border_args.get('top', cell.border.top),
                    bottom=border_args.get('bottom', cell.border.bottom)
                )
        # --- Function to create the Base Reagent Prep table at a given location for a given reagent ---
        def create_base_reagent_prep_table(ws_prep, ws_comp, reagent, top_left_row, top_left_col):
            """
            Create the Base Reagent Prep table for a given reagent at the specified top-left cell (row, col).
            """


            base_reagent = donor_id_dict.get(reagent)

            prep_headers = [
                base_reagent,
                'Total Volume (uL)',
                'Amount Needed (nmol)'
            ]

            from openpyxl.styles import Border, Side, Font, Alignment
            right_border = Border(right=Side(border_style="thin"))
            # Write headers
            for i, header in enumerate(prep_headers):
                ws_prep.cell(row=top_left_row + i, column=top_left_col).value = header
                ws_prep.cell(row=top_left_row + i, column=top_left_col).font = Font(bold=True)
                ws_prep.cell(row=top_left_row + i, column=top_left_col).border = right_border
                ws_prep.cell(row=top_left_row + i, column=top_left_col).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            
            # Add header for Volume (uL) above the third column
            ws_prep.cell(row=top_left_row, column=top_left_col+2).value = 'Volume (uL)'
            ws_prep.cell(row=top_left_row, column=top_left_col+2).font = Font(bold=True)
            ws_prep.cell(row=top_left_row, column=top_left_col+2).alignment = Alignment(horizontal='center', vertical='center')
            
            # Remove any old header above (only if row > 1)
            if top_left_row > 1:
                ws_prep.cell(row=top_left_row-1, column=top_left_col+2).value = None

            # Write values (formulas) for the reagent
            found = False
            for comp_row in range(2, ws_comp.max_row+1):
                comp_name = ws_comp.cell(row=comp_row, column=start_column_index + 1).value
                if comp_name == reagent:
                    from openpyxl.utils import get_column_letter
                    # For column B (originally 2), add start_column_index
                    col_B = get_column_letter(1 + start_column_index)
                    col_F = get_column_letter(5 + start_column_index)
                    col_H = get_column_letter(7 + start_column_index)
                    # Write the component name in the first column of the prep table
                    ws_prep.cell(row=top_left_row, column=top_left_col+1).value = f"='Base Reagent Cal.'!{col_B}{comp_row}"
                    ws_prep.cell(row=top_left_row, column=top_left_col+1).alignment = Alignment(horizontal='center', vertical='center')
                    ws_prep.cell(row=top_left_row, column=top_left_col+1).font = Font(bold=True)
                    ws_prep.cell(row=top_left_row+1, column=top_left_col+1).value = f"='Base Reagent Cal.'!{col_F}{comp_row}"
                    ws_prep.cell(row=top_left_row+2, column=top_left_col+1).value = f"='Base Reagent Cal.'!{col_H}{comp_row}"
                    found = True
                    break
            if not found:
                ws_prep.cell(row=top_left_row, column=top_left_col+1).value = reagent
                ws_prep.cell(row=top_left_row, column=top_left_col+1).alignment = Alignment(horizontal='center', vertical='center')
                ws_prep.cell(row=top_left_row, column=top_left_col+1).font = Font(bold=True)
                ws_prep.cell(row=top_left_row+1, column=top_left_col+1).value = ''
                ws_prep.cell(row=top_left_row+2, column=top_left_col+1).value = ''
            # Add 6 new rows for stock conc. and stock volume for 1, 2, 3, with default value 0
            extra_labels = [
                'Stock Conc. 1 (uM)', 'Stock Volume 1 (uL)',
                'Stock Conc. 2 (uM)', 'Stock Volume 2 (uL)',
                'Stock Conc. 3 (uM)', 'Stock Volume 3 (uL)'
            ]
            for i, label in enumerate(extra_labels):
                row = top_left_row + 3 + i
                ws_prep.cell(row=row, column=top_left_col).value = label
                ws_prep.cell(row=row, column=top_left_col).font = Font(bold=True)
                ws_prep.cell(row=row, column=top_left_col).border = right_border
                ws_prep.cell(row=row, column=top_left_col).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                ws_prep.cell(row=row, column=top_left_col+1).value = 0
            # Write formulas for stock volumes in the volume column (col 3)
            ws_prep.cell(row=top_left_row+4, column=top_left_col+2).value = f"=IF(AND({chr(65+top_left_col)}{top_left_row+3}>0,{chr(65+top_left_col)}{top_left_row+2}>0),MIN({chr(65+top_left_col)}{top_left_row+4},ROUND({chr(65+top_left_col)}{top_left_row+2}*1000/{chr(65+top_left_col)}{top_left_row+3},2)),0)"
            ws_prep.cell(row=top_left_row+6, column=top_left_col+2).value = f"=IF(AND({chr(65+top_left_col)}{top_left_row+5}>0,MAX(0,{chr(65+top_left_col)}{top_left_row+2}-{chr(65+top_left_col+1)}{top_left_row+4}*{chr(65+top_left_col)}{top_left_row+3}/1000)>0),MIN({chr(65+top_left_col)}{top_left_row+6},ROUND(MAX(0,{chr(65+top_left_col)}{top_left_row+2}-{chr(65+top_left_col+1)}{top_left_row+4}*{chr(65+top_left_col)}{top_left_row+3}/1000)*1000/{chr(65+top_left_col)}{top_left_row+5},2)),0)"
            ws_prep.cell(row=top_left_row+8, column=top_left_col+2).value = f"=IF(AND({chr(65+top_left_col)}{top_left_row+7}>0,MAX(0,{chr(65+top_left_col)}{top_left_row+2}-{chr(65+top_left_col+1)}{top_left_row+4}*{chr(65+top_left_col)}{top_left_row+3}/1000-{chr(65+top_left_col+1)}{top_left_row+6}*{chr(65+top_left_col)}{top_left_row+5}/1000)>0),MIN({chr(65+top_left_col+2)}{top_left_row+8},ROUND(MAX(0,{chr(65+top_left_col)}{top_left_row+2}-{chr(65+top_left_col+1)}{top_left_row+4}*{chr(65+top_left_col)}{top_left_row+3}/1000-{chr(65+top_left_col+1)}{top_left_row+6}*{chr(65+top_left_col)}{top_left_row+5}/1000)*1000/{chr(65+top_left_col)}{top_left_row+7},2)),0)"
            # At the bottom, add a cell for 'Required nmol needed' and formula
            ws_prep.cell(row=top_left_row+9, column=top_left_col).value = 'Required nmol needed'
            ws_prep.cell(row=top_left_row+9, column=top_left_col).font = Font(bold=True)
            ws_prep.cell(row=top_left_row+9, column=top_left_col).border = right_border
            ws_prep.cell(row=top_left_row+9, column=top_left_col).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            ws_prep.cell(row=top_left_row+9, column=top_left_col+1).value = f"=ROUND(MAX(0,{chr(65+top_left_col)}{top_left_row+2}-{chr(65+top_left_col+1)}{top_left_row+4}*{chr(65+top_left_col)}{top_left_row+3}/1000-{chr(65+top_left_col+1)}{top_left_row+6}*{chr(65+top_left_col)}{top_left_row+5}/1000-{chr(65+top_left_col+1)}{top_left_row+8}*{chr(65+top_left_col)}{top_left_row+7}/1000),2)"
            # Add a border between the second and third column for all relevant rows
            mid_right_border = Border(right=Side(border_style="thin"))
            for i in range(0, 10):
                ws_prep.cell(row=top_left_row + i, column=top_left_col+1).border = mid_right_border
            # Auto-fit columns for this table: set all columns to the width of the first column
            first_col_letter = ws_prep.cell(row=top_left_row, column=top_left_col).column_letter
            first_col_width = max(len(str(ws_prep.cell(row=top_left_row + i, column=top_left_col).value)) for i in range(0, 10)) + 2
            for col in range(top_left_col, top_left_col+3):
                col_letter = ws_prep.cell(row=top_left_row, column=col).column_letter
                ws_prep.column_dimensions[col_letter].width = first_col_width
            # Place the Additional Reagent section below the main table, with only one blank row for spacing
            add_reagent_start_row = top_left_row + 10  # 10 rows for main table
            ws_prep.cell(row=add_reagent_start_row, column=top_left_col).value = None
            # Now add the additional reagents table below the single blank row
            add_table_row = add_reagent_start_row + 1
            if reagent in ['Ubc13/Mms2', 'gp78/Ube2g2', 'Ube2K']:
                extra_reagents = ['hUba1 (uM)', 'TCEP (mM)', 'ATP/Mg2+ (mM)']
            else:
                extra_reagents = ['TCEP (mM)', 'ATP/Mg2+ (mM)']
            extra_col = top_left_col
            ws_prep.cell(row=add_table_row, column=extra_col).value = 'Reagent'
            ws_prep.cell(row=add_table_row, column=extra_col+1).value = 'Final Conc.'
            ws_prep.cell(row=add_table_row, column=extra_col+2).value = 'Volume (uL)'
            # Style header row for second table
            for j, label in enumerate(['Reagent', 'Final Conc.', 'Volume (uL)']):
                cell = ws_prep.cell(row=add_table_row, column=extra_col+j)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # Fill in second table and style cells
            for i, reagent_name in enumerate(extra_reagents):
                for j in range(3):
                    cell = ws_prep.cell(row=add_table_row+1+i, column=extra_col+j)
                    # Only the first column (Reagent) is bold, others are not
                    if j == 0:
                        cell.font = Font(bold=True)
                    else:
                        cell.font = Font(bold=False)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                ws_prep.cell(row=add_table_row+1+i, column=extra_col).value = reagent_name
                # Final Conc.
                if reagent_name == 'hUba1 (uM)':
                    ws_prep.cell(row=add_table_row+1+i, column=extra_col+1).value = default_final_concentrations.get('hUba1', '')
                elif reagent_name == 'TCEP (mM)':
                    ws_prep.cell(row=add_table_row+1+i, column=extra_col+1).value = default_final_concentrations.get('TCEP', '')
                elif reagent_name == 'ATP/Mg2+ (mM)':
                    ws_prep.cell(row=add_table_row+1+i, column=extra_col+1).value = default_final_concentrations.get('ATP/Mg2+', '')
                # Set Volume (uL) formula
                comp_vol_cell = ws_prep.cell(row=top_left_row+1, column=top_left_col+1).coordinate
                if reagent_name == 'hUba1 (uM)':
                    init_conc_cell = ws_prep.cell(row=3, column=3).coordinate  # C3
                elif reagent_name == 'TCEP (mM)':
                    init_conc_cell = ws_prep.cell(row=4, column=3).coordinate  # C4
                elif reagent_name == 'ATP/Mg2+ (mM)':
                    init_conc_cell = ws_prep.cell(row=5, column=3).coordinate  # C5
                else:
                    init_conc_cell = None
                final_conc_cell = ws_prep.cell(row=add_table_row+1+i, column=extra_col+1).coordinate
                if init_conc_cell:
                    if reagent_name == 'hUba1 (uM)':
                        ws_prep.cell(row=add_table_row+1+i, column=extra_col+2).value = f"=IF({init_conc_cell}>0,({final_conc_cell}*2*{comp_vol_cell})/{init_conc_cell},0)"
                    else:
                        ws_prep.cell(row=add_table_row+1+i, column=extra_col+2).value = f"=IF({init_conc_cell}>0,({final_conc_cell}*{comp_vol_cell})/{init_conc_cell},0)"
                else:
                    ws_prep.cell(row=add_table_row+1+i, column=extra_col+2).value = ''
            # Add HEPES buffer row
            hepes_row = add_table_row+1+len(extra_reagents)
            for j in range(3):
                cell = ws_prep.cell(row=hepes_row, column=extra_col+j)
                # Only the first column (Reagent) is bold, others are not
                if j == 0:
                    cell.font = Font(bold=True)
                else:
                    cell.font = Font(bold=False)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_prep.cell(row=hepes_row, column=extra_col).value = 'HEPES buffer'
            ws_prep.cell(row=hepes_row, column=extra_col+1).value = ''
            # Build formula for HEPES buffer volume
            # Base Reagent volume cell
            comp_vol_cell = ws_prep.cell(row=top_left_row+1, column=top_left_col+1).coordinate
            # Collect all other volume cells in both tables
            volume_cells = []
            # Main table: stock volumes (every other row, starting from top_left_row+4, col=top_left_col+2, 3 stock volumes)
            for k in [4,6,8]:
                volume_cells.append(ws_prep.cell(row=top_left_row+k, column=top_left_col+2).coordinate)
            # Additional reagents table: hUba1, TCEP, ATP/Mg2+ (rows add_table_row+1 to add_table_row+len(extra_reagents), col=extra_col+2)
            for k in range(len(extra_reagents)):
                volume_cells.append(ws_prep.cell(row=add_table_row+1+k, column=extra_col+2).coordinate)
            sum_formula = '+'.join(volume_cells)
            ws_prep.cell(row=hepes_row, column=extra_col+2).value = f"={comp_vol_cell}-({sum_formula})"
            # --- Draw a single, unbroken medium border around both tables as one block ---
            from openpyxl.styles import Border, Side
            outer_side = Side(border_style="medium")  # less bold than thick
            inner_side = Side(border_style="thin")

            # Add a light bold (medium) border under the title line of the component calcualtion table & additional reagents table
            for row_num in [top_left_row, add_table_row, add_table_row -1]:
                for j in range(3):
                    cell = ws_prep.cell(row=row_num, column=extra_col + j)
                    cell.border = Border(
                        top=cell.border.top,
                        left=cell.border.left,
                        right=cell.border.right,
                        bottom=Side(border_style="medium")
                    )

            # Rectangle: from top_left_row, top_left_col to bottom row/col of second table
            # Bottom border should be directly under HEPES buffer (not after the extra blank row)
            total_rows = (10 + 1 + 1 + len(extra_reagents) + 1)  # main table + blank + header + extra reagents + HEPES
            start_row = top_left_row
            end_row = hepes_row  # last row with content is HEPES buffer
            start_col = top_left_col
            end_col = top_left_col + 2
            # Add a blank row after HEPES buffer for visual separation
            for c in range(start_col, end_col+1):
                ws_prep.cell(row=hepes_row+1, column=c).value = None
            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    cell = ws_prep.cell(row=r, column=c)
                    border_args = {}
                    # Outer border
                    if r == start_row:
                        border_args['top'] = outer_side
                    if r == end_row:
                        border_args['bottom'] = outer_side
                    if c == start_col:
                        border_args['left'] = outer_side
                    if c == end_col:
                        border_args['right'] = outer_side
                    # For the bottom (additional reagents) table, add thin vertical lines between columns (now up to HEPES buffer)
                    if r >= add_table_row and r <= end_row:
                        if c == start_col + 1:
                            border_args['left'] = inner_side
                        if c == start_col + 1:
                            border_args['right'] = inner_side
                    if border_args:
                        existing = cell.border
                        cell.border = Border(
                            left=border_args.get('left', existing.left),
                            right=border_args.get('right', existing.right),
                            top=border_args.get('top', existing.top),
                            bottom=border_args.get('bottom', existing.bottom)
                        )
        # Create the Base Reagent Prep table for each reagent at specified locations
        # Only create the Base Reagent Prep table for 'Ubc13/Mms2' at (5, 4)
        create_base_reagent_prep_table(ws_prep, ws_comp, 'Ubc13/Mms2', 2, 5)
        create_base_reagent_prep_table(ws_prep, ws_comp, 'gp78/Ube2g2', 20, 5)
        create_base_reagent_prep_table(ws_prep, ws_comp, 'Ube2K', 38, 5)

        # Create the Base Reagent Prep table for the donors
        create_base_reagent_prep_table(ws_prep, ws_comp, 'ubi_ubq_1_K48_ABOC_K63_ABOC', 74, 10) 
        create_base_reagent_prep_table(ws_prep, ws_comp, 'ubi_ubq_1_K48_SMAC_K63_ABOC', 20, 10)  
        create_base_reagent_prep_table(ws_prep, ws_comp, 'ubi_ubq_1_K48_ABOC_K63_SMAC', 56, 10) 
        create_base_reagent_prep_table(ws_prep, ws_comp, 'ubi_ubq_1_K48_SMAC', 2, 10) 
        create_base_reagent_prep_table(ws_prep, ws_comp, 'ubi_ubq_1_K63_SMAC', 38, 10) 
        
        # Widen the 11th column (column 'K') of the 'Base Reagent Prep' worksheet for better readability
        ws_prep = wb['Base Reagent Prep']
        ws_prep.column_dimensions['K'].width = 28

        # --- New: Add Acceptor Dimer Cal. worksheet if acceptors_96_counts is provided ---
        def add_acceptor_counts_sheet(wb, acceptors_96_counts):
            """
            Add a worksheet displaying acceptor counts (acceptors_96_counts dict) to the workbook.
            Also display dimer_encoded_dictionary as a table using pandas, starting at D2.
            """

            # Check if acceptors_96_counts is provided and not empty
            ws_acceptor = wb.create_sheet('Acceptor Dimer Cal.')
            
            column_start = 2  # Start at column B

            # --- Display merged dimer_encoded_dictionary and acceptors_96_counts as a table using pandas, starting at D2 ---
            if dimer_encoded_dictionary and acceptors_96_counts:
                import pandas as pd
                # dimer_encoded_dictionary: {encoded_value: dimer_key}
                # acceptors_96_counts: {encoded_value: count}
                merged = []
                for dimer_name, encoded_value in dimer_encoded_dictionary.items():
                    count = acceptors_96_counts.get(encoded_value, 0)
                    merged.append({'Encoded Value': encoded_value, 'Acceptor': dimer_name, 'Count': count})

                # Merge acceptor_description_dict into merged based on Encoded Value <-> No.
                # Build a lookup from number to description dict
                desc_by_number = {}
                for desc in acceptor_description_dict:
                    try:
                        num = int(desc["No."].split()[-1])
                        desc_by_number[num] = desc
                    except Exception:
                        continue
                for row in merged:
                    encoded_val = row['Encoded Value']
                    desc = desc_by_number.get(encoded_val)
                    if desc:
                        for k, v in desc.items():
                            if k != "No." and k not in row:
                                row[k] = v

                merged_df = pd.DataFrame(merged)
                # Add 'Acceptor No.' column from acceptor_description_dict if possible
                acceptor_no_map = {int(desc["No."].split()[-1]): desc["No."] for desc in acceptor_description_dict}
                merged_df.insert(1, 'Acceptor No.', merged_df['Encoded Value'].map(acceptor_no_map))
                # Remove the 'Acceptor' column if present
                if 'Acceptor' in merged_df.columns:
                    merged_df = merged_df.drop(columns='Acceptor')
                # Move 'Count' column to the end
                if 'Count' in merged_df.columns:
                    count_col = merged_df.pop('Count')
                    merged_df['Count'] = count_col
                # Add 'nmol / well' and 'Total nmol' columns after 'Count'
                nmol_per_well_default = 10
                if 'Count' in merged_df.columns:
                    count_idx = merged_df.columns.get_loc('Count')
                    merged_df.insert(count_idx + 1, 'nmol / well', nmol_per_well_default)
                    merged_df.insert(count_idx + 2, 'Total nmol', '')  # Placeholder for Excel formula
                # Add 'Stock Conc. (uM)', 'Volume (uL) / well', and 'Total Volume needed (uL)' columns at the end
                stock_conc_default = 20
                merged_df['Stock Conc. (uM)'] = stock_conc_default
                merged_df['Volume (uL) / well'] = ''  # Placeholder for Excel formula
                merged_df['Total Volume needed (uL)'] = ''  # Placeholder for Excel formula

                # Write headers (including new columns)
                for col_idx, col_name in enumerate(merged_df.columns, start=column_start):
                    ws_acceptor.cell(row=2, column=col_idx).value = col_name
                    ws_acceptor.cell(row=2, column=col_idx).font = header_font
                    ws_acceptor.cell(row=2, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                # Add a bold line (medium border) under the title line
                from openpyxl.styles import Border, Side
                for col_idx in range(column_start, column_start + len(merged_df.columns)):
                    cell = ws_acceptor.cell(row=2, column=col_idx)
                    cell.border = Border(
                        top=cell.border.top,
                        left=cell.border.left,
                        right=cell.border.right,
                        bottom=Side(border_style="medium")
                    )
                # Write data and set Excel formulas for new columns
                for row_idx, row in enumerate(merged_df.itertuples(index=False), start=3):
                    for col_idx, value in enumerate(row, start=column_start):
                        col_name = merged_df.columns[col_idx - column_start]
                        if col_name == 'Total nmol':
                            count_col_letter = ws_acceptor.cell(row=2, column=col_idx-2).column_letter
                            nmol_col_letter = ws_acceptor.cell(row=2, column=col_idx-1).column_letter
                            ws_acceptor.cell(row=row_idx, column=col_idx).value = f"={count_col_letter}{row_idx}*{nmol_col_letter}{row_idx}"
                        elif col_name == 'Volume (uL) / well':
                            nmol_col_letter = ws_acceptor.cell(row=2, column=col_idx-3).column_letter
                            stock_col_letter = ws_acceptor.cell(row=2, column=col_idx-1).column_letter
                            ws_acceptor.cell(row=row_idx, column=col_idx).value = f"=ROUND({nmol_col_letter}{row_idx}/{stock_col_letter}{row_idx}*1000,2)"
                        elif col_name == 'Total Volume needed (uL)':
                            total_nmol_col_letter = ws_acceptor.cell(row=2, column=col_idx-3).column_letter
                            stock_col_letter = ws_acceptor.cell(row=2, column=col_idx-2).column_letter
                            ws_acceptor.cell(row=row_idx, column=col_idx).value = f"=ROUND({total_nmol_col_letter}{row_idx}/{stock_col_letter}{row_idx}*1000,2)"
                        else:
                            ws_acceptor.cell(row=row_idx, column=col_idx).value = value
                        ws_acceptor.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                # Auto-fit all columns to the width of their header text
                for col_idx, col_name in enumerate(merged_df.columns, start=column_start):
                    col_letter = ws_acceptor.cell(row=2, column=col_idx).column_letter
                    header_length = len(str(col_name))
                    ws_acceptor.column_dimensions[col_letter].width = header_length + 2

        # At the end of create_combined_xlsx, add acceptor counts sheet if provided
        if acceptors_96_counts:
            add_acceptor_counts_sheet(wb, acceptors_96_counts)
        

        # When done, save back to BytesIO for FastAPI
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)
        
        return final_output
    
    return inner_create_xlsx_bytes(enzyme_donor_counts, component_counts)

def create_opentrons_file_bytes(output_dict):
    
    """
    Create an Opentrons protocol file from the output dictionary.
    The output_dict should contain 'deprots_96' DataFrame and 'deprot_encoded_dictionary'.
    Returns a bytes object containing the Opentrons protocol code.
    """
    import io
    
    deprots_96_df = (output_dict.get('deprots_96'))   
    reaction_sets = []
    cols = deprots_96_df.shape[1]
    for i in range(0, cols, 2):
        subset = deprots_96_df.iloc[:, i:i+2].values.T.flatten()
        ones_and_twos = [int(v) for v in subset if v in (1, 2)]
        reaction_sets.append(ones_and_twos)
    reaction_sets = [s for s in reaction_sets if len(s) > 0]
    inv_deprot_dict = {v: k for k, v in output_dict.get('deprot_encoded_dictionary', {}).items()}
    reaction_sets_named = [[inv_deprot_dict.get(val, val) for val in s] for s in reaction_sets]

    opentrons_code = '''from typing import List\nfrom opentrons import protocol_api\nfrom opentrons.protocol_api import InstrumentContext, labware\nfrom io import StringIO\nimport time\n\nmetadata = {\n    "ctxName": "OTflex_dimer formation",\n    "author": "Bode Lab",\n    "description": "Original Author of Opentrons code: Sohei Majima",\n}\n\nrequirements = {"robotType": "Flex", "apiLevel": "2.18"}\n\n##############################\n\n'''
    opentrons_code += 'data = ' + repr(reaction_sets_named) + '\n\n'
    opentrons_code += '''##############################\n\nisolation_time = 1.5 #should be 5. Only for testing.\ntransfer_amount = 200\ntemperature = 40\nspeed = 700\n\n# ...rest of protocol code...\n'''
    
    opentrons_code = '''
from typing import List
from opentrons import protocol_api
from opentrons.protocol_api import InstrumentContext, labware
from io import StringIO
import time

metadata = {
    "ctxName": "OTflex_dimer formation",
    "author": "Bode Lab",
    "description": "Original Author of Opentrons code: Sohei Majima",
}

requirements = {"robotType": "Flex", "apiLevel": "2.18"}

##############################\n\n'''
# Copy the data array from the previous code snippet
    opentrons_code += 'data = ' + repr(reaction_sets_named)
    opentrons_code += '''

##############################

isolation_time = 1.5 #should be 5. Only for testing.
transfer_amount = 200
temperature = 40
speed = 700


def add_parameters(parameters):

    parameters.add_int(
        variable_name="ub_len",
        display_name="ub_length",
        description="Are you synthesizing tetramer?",
        default=5,
        choices=[
            {"display_name": "tetramer", "value": 4},
            {"display_name": "pentamer", "value": 5},
            {"display_name": "dimer", "value": 2},
        ],
    )

    parameters.add_int(
        variable_name="scale",
        display_name="scale",
        description="What's the scale of the reaction in µl?",
        default=200,
        minimum=100,
        maximum=600
        )
    
    parameters.add_int(
        variable_name="acceptor",
        display_name="acceptor",
        description="What is your acceptor?",
        default=2,
        choices=[
            {"display_name": "dimer", "value": 2},
            {"display_name": "monomer", "value": 1},
        ],
    )

    parameters.add_int(
        variable_name="test",
        display_name="test",
        description="Is this a test run (short reaction time)?",
        default=0,
        choices=[
            {"display_name": "yes", "value": 1},
            {"display_name": "no", "value": 0},
        ],
    )

    parameters.add_int(
        variable_name="shaker",
        display_name="shaker speed",
        description="What's the speed of the shaker for wash?",
        default=800,
        minimum=100,
        maximum=1000
        )

            
##########################


def run(ctx):

    global num_sample
    global ub_len
    global scale
    global acceptor
    global test
    global shaker

    ub_len = ctx.params.ub_len
    scale = ctx.params.scale   
    acceptor = ctx.params.acceptor
    test = ctx.params.test
    shaker = ctx.params.shaker

    num_sample = '''
    opentrons_code += str(int(len(reaction_sets_named[0])))
    opentrons_code += '''

    if test == 1:
        reaction_time = 0.5
        deprotection_time = 0.5
    else:
        reaction_time = 600 #stick to it if you want to run it with in 3 days
        deprotection_time = 120


    # load labware
    heater_shaker = ctx.load_module('heaterShakerModuleV1', 'C1')
    hs_adapter = heater_shaker.load_adapter('opentrons_96_deep_well_adapter')
    magnetic_module = ctx.load_module("magneticBlockV1", 'D1')
    mm_adapter = magnetic_module.load_adapter('opentrons_96_deep_well_adapter')
    reaction_plate = mm_adapter.load_labware('nest_96_wellplate_2ml_deep')
    tube_rack = ctx.load_labware('opentrons_6_tuberack_falcon_50ml_conical', "D2")
    tiprack_200 = ctx.load_labware("opentrons_flex_96_tiprack_200ul", "B3")
    tiprack_200_2 = ctx.load_labware("opentrons_flex_96_tiprack_200ul", "C3")
    tiprack_200_3 = ctx.load_labware("opentrons_flex_96_tiprack_200ul", "D3")
    tiprack_200_4 = ctx.load_labware("opentrons_flex_96_tiprack_200ul", "A2")
    tiprack_200_5 = ctx.load_labware("opentrons_flex_96_tiprack_200ul", "B1")
    paradox_plate = ctx.load_labware('inhecoparadoxplate_96_tuberack_1000ul', "A1")

    trash = ctx.load_trash_bin('A3')

    left_pipette = ctx.load_instrument(
        'flex_1channel_1000', 'left', tip_racks=[tiprack_200, tiprack_200_2, tiprack_200_3, tiprack_200_4, tiprack_200_5])
    right_pipette = ctx.load_instrument(
        'flex_8channel_1000', 'right', tip_racks=[tiprack_200, tiprack_200_2, tiprack_200_3, tiprack_200_4, tiprack_200_5])

    default_rate = 300
    left_pipette.flow_rate.aspirate = default_rate
    right_pipette.flow_rate.dispense = default_rate

    
    wash_stock = ctx.load_labware('nest_12_reservoir_15ml', 'B2', 'wash buffer')
    depro_wash_stock = ctx.load_labware('nest_12_reservoir_15ml', 'C2', 'deprotection wash buffer')

    ####liquid
    wash_def = ctx.define_liquid(name="WASH", description="Wash Buffer", display_color="#9ACECB")  

    depro_wash_def = ctx.define_liquid(name="DEPROTECTION_WASH", description="Deprotection Wash Buffer", display_color="#808080")  

    wash_stock["A1"].load_liquid(liquid=wash_def, volume=15000) 
    depro_wash_stock["A1"].load_liquid(liquid=depro_wash_def, volume=15000) 

    ## supporting functions

    number_of_ubiquitin = int(num_sample)
    number_of_lane = (number_of_ubiquitin + 8 - 1) // 8

    def min(x, y):
        return x if x < y else y

    def sample_to_well_positions(number_of_ubiquitin):
        rows = "ABCDEFGH"
        columns = range(1, 13)

        well_positions = []

        if number_of_ubiquitin < 1 or number_of_ubiquitin > 96:
            raise ValueError("Sample number must be between 1 and 96")

        for i in range(0, number_of_ubiquitin):
            row = rows[i % 8]
            column = columns[i // 8]
            well_positions.append(f"{row}{column}")
    
        return well_positions

    well_position = sample_to_well_positions(number_of_ubiquitin)

    #########
    def reaction(source_column: tuple = (1, 2)):
        ctx.comment("This is the beginning of ubiquitylation reaction.")
        right_pipette.flow_rate.aspirate = 50
        right_pipette.flow_rate.dispense = 10
        
        #add reaction mixtures to each lanes
        i = 1
        while i <= source_column[0]:
            right_pipette.pick_up_tip()
            right_pipette.transfer(
                min(scale, 200),
                paradox_plate[f'A{i + (source_column[1] - 1)*number_of_lane}'],  #source
                reaction_plate[f'A{i}'].bottom(3),  #reaction well
                blow_out=True,
                blowout_location='destination well',
                new_tip='never'
            )
            #double if scale is more than 300 µl

            if scale > 200:
                right_pipette.transfer(
                    scale - 200,
                    paradox_plate[f'A{i + (source_column[1] - 1)*number_of_lane}'],  
                    reaction_plate[f'A{i}'].bottom(3),
                    #This can be somewhere else. How to determine?
                    blow_out=True,
                    blowout_location='destination well',
                    new_tip='never'
                )
            right_pipette.drop_tip()
            i += 1
        
        # adding lid
        ctx.move_labware(labware = reaction_plate,
                            new_location=hs_adapter,
                            use_gripper=True
                            )


        heater_shaker.close_labware_latch()
        ctx.pause('Seal the 96 well plate')

        heater_shaker.set_target_temperature(temperature)
        heater_shaker.set_and_wait_for_shake_speed(speed)
        ctx.delay(minutes=reaction_time)  # 16 hours reaction
        heater_shaker.deactivate_heater()
        heater_shaker.deactivate_shaker()

    
    def wash_reaction(source_column: tuple = (1, 2)):
        #removal of the solution with right pipette
        #source column[0] is number of lane, source column[1] is number of ubiquitylation, (dimer = 1, trimer = 2, tetramer = 3, pentamer = 4)
        ctx.comment("This is the beginning of wash.")
        #Removal of reaction mixture
        right_pipette.flow_rate.aspirate = 10
        right_pipette.flow_rate.dispense = 200
        heater_shaker.open_labware_latch()
        # removing lid
        ctx.pause('Remove the seal of the 96 well plate and centrifuge the plate')
        #moving from heater shaker to magnetic module
        ctx.move_labware(labware = reaction_plate,
                            new_location=mm_adapter,
                            use_gripper=True
                            )
        ctx.delay(minutes=isolation_time)
        j = 1
        right_pipette.pick_up_tip()
        while j <= source_column[0]:
            right_pipette.transfer(
                scale,
                reaction_plate[f'A{j}'],
                reaction_plate[f'A{j + (source_column[1])*number_of_lane}'], 
                new_tip='never'
            )
            j += 1
        right_pipette.drop_tip()
        
    ###wash for 3 times
        for i in range(2):
            right_pipette.flow_rate.aspirate = 200
            right_pipette.flow_rate.dispense = 10
            right_pipette.pick_up_tip()
            k = 0
            while k <= number_of_lane - 1:
                #left_pipette.flow_rate.aspirate = 50
                #left_pipette.pick_up_tip()
                right_pipette.transfer(
                    transfer_amount,
                    wash_stock['A1'],
                    reaction_plate[f'A{k+1}'].bottom(3),
                    new_tip='never'
                )
                #left_pipette.mix(5, 200, magnetic_plate[f'{well_position[k]}'])
                #left_pipette.drop_tip()
                k += 1
            
            right_pipette.drop_tip()

            ctx.move_labware(labware = reaction_plate,
                            new_location=hs_adapter,
                            use_gripper=True
                            )
            heater_shaker.close_labware_latch()
            heater_shaker.set_and_wait_for_shake_speed(shaker)
            ctx.delay(minutes=0.5)
            heater_shaker.deactivate_shaker()    
            heater_shaker.open_labware_latch()
            ctx.move_labware(labware = reaction_plate,
                            new_location=mm_adapter,
                            use_gripper=True
                            )
            ctx.delay(minutes=isolation_time)

                # wash removal using 8 channel pipettes
            l = 1
            right_pipette.pick_up_tip()
            while l <= number_of_lane:
                right_pipette.flow_rate.aspirate = 10
                right_pipette.flow_rate.dispense = 200
                right_pipette.aspirate(transfer_amount, reaction_plate[f'A{l}'])
                right_pipette.dispense(transfer_amount, trash)
                #right_pipette.transfer(
                    #transfer_amount,
                    #reaction_plate[f'A{l}'],
                    #trash,
                    #new_tip='never'
                #)
                l += 1
            right_pipette.drop_tip()
        
        #left_pipette.drop_tip()
        #right_pipette.drop_tip()
        ctx.comment("This is the end of wash.")

    def deprotection(source_column = (1)):
        heater_shaker.open_labware_latch()
        #left_pipette.pick_up_tip()
        left_pipette.flow_rate.aspirate = 200
        left_pipette.flow_rate.dispense = 10
        
        for i in range(len(data[0])):
            if data[source_column][i] == 'SMAC_deprot':
                left_pipette.transfer(
                transfer_amount,
                tube_rack['A1'], #change later
                reaction_plate[f'{well_position[i]}'].bottom(3),
                blow_out=True,
                blowout_location='destination well'
                )
            elif data[source_column][i] == 'FAKE_deprot':
                left_pipette.transfer(
                transfer_amount,
                tube_rack['A2'], #change later
                reaction_plate[f'{well_position[i]}'].bottom(3),
                blow_out=True,
                blowout_location='destination well'
                )
            else:
                print('Error')

        #left_pipette.drop_tip()
        ctx.move_labware(labware = reaction_plate,
                            new_location=hs_adapter,
                            use_gripper=True
                            )
        heater_shaker.close_labware_latch()
        heater_shaker.set_target_temperature(temperature)
        heater_shaker.set_and_wait_for_shake_speed(speed)
        ctx.delay(minutes=deprotection_time)  # 60 min
        heater_shaker.deactivate_heater()
        heater_shaker.deactivate_shaker()
        heater_shaker.open_labware_latch()

    def wash_deprotection(source_column: tuple = (1)):
        #removal of the solution with right pipette
        #source column[0] is number of lane
        
        #Removal of reaction mixture
        ctx.comment("This is the beginning of wash of PLP solution.")
        ctx.pause('Centrifuge the plate')
        right_pipette.flow_rate.aspirate = 10
        right_pipette.flow_rate.dispense = 200
        ctx.move_labware(labware = reaction_plate,
                            new_location=mm_adapter,
                            use_gripper=True
                            )
        ctx.delay(minutes=isolation_time)
        j = 1
        right_pipette.pick_up_tip()

        while j <= source_column:
            right_pipette.aspirate(transfer_amount, reaction_plate[f'A{j}'])
            right_pipette.dispense(transfer_amount, trash)
            j += 1
        right_pipette.drop_tip()

        #right_pipette.pick_up_tip()

        ### wash 2 times
        for i in range(2):
            k = 0
            right_pipette.pick_up_tip()
            while k <= number_of_lane - 1:
                right_pipette.flow_rate.aspirate = 200
                right_pipette.flow_rate.dispense = 10
                #left_pipette.pick_up_tip()
                right_pipette.transfer(
                    transfer_amount,
                    depro_wash_stock['A1'],
                    reaction_plate[f'A{k+1}'].bottom(3),
                    new_tip='never'
                )
            #left_pipette.mix(5, 200, magnetic_plate[f'{well_position[k]}'])
            #left_pipette.drop_tip()
                k += 1
            right_pipette.drop_tip()

            ctx.move_labware(labware = reaction_plate,
                                new_location=hs_adapter,
                                use_gripper=True
                                )
            heater_shaker.close_labware_latch()
            
            heater_shaker.set_and_wait_for_shake_speed(shaker)
            ctx.delay(minutes=0.5)
            heater_shaker.deactivate_shaker()    
            heater_shaker.open_labware_latch()
            ctx.move_labware(labware = reaction_plate,
                                new_location=mm_adapter,
                                use_gripper=True
                                )
            ctx.delay(minutes=isolation_time)

            # wash removal using 8 channel pipettes
            l = 1
            right_pipette.pick_up_tip()
            while l <= number_of_lane:
                right_pipette.flow_rate.aspirate = 10
                right_pipette.flow_rate.dispense = 200
                right_pipette.aspirate(transfer_amount, reaction_plate[f'A{l}'])
                right_pipette.dispense(transfer_amount, trash)
                l += 1
            right_pipette.drop_tip()

        ctx.comment("Equilibration of resin to buffer")
        #right_pipette.pick_up_tip()
        ### wash 2 times
        for i in range(2):
            right_pipette.pick_up_tip()
            k = 0
            while k <= number_of_lane - 1:
                right_pipette.flow_rate.aspirate = 200
                right_pipette.flow_rate.dispense = 10
                #left_pipette.pick_up_tip()
                right_pipette.transfer(
                    transfer_amount,
                    wash_stock['A1'],
                    reaction_plate[f'A{k+1}'].bottom(3),
                    new_tip='never'
                )
            #left_pipette.mix(5, 200, magnetic_plate[f'{well_position[k]}'])
            #left_pipette.drop_tip()
                k += 1
            right_pipette.drop_tip()

            ctx.move_labware(labware = reaction_plate,
                                new_location=hs_adapter,
                                use_gripper=True
                                )
            heater_shaker.close_labware_latch()
            
            heater_shaker.set_and_wait_for_shake_speed(shaker)
            ctx.delay(minutes=0.5)
            heater_shaker.deactivate_shaker()    
            heater_shaker.open_labware_latch()
            ctx.move_labware(labware = reaction_plate,
                                new_location=mm_adapter,
                                use_gripper=True
                                )
            ctx.delay(minutes=isolation_time)

            # wash removal using 8 channel pipettes
            l = 1
            right_pipette.pick_up_tip()
            while l <= number_of_lane:
                right_pipette.flow_rate.aspirate = 10
                right_pipette.flow_rate.dispense = 200
                right_pipette.aspirate(transfer_amount, reaction_plate[f'A{l}'])
                right_pipette.dispense(transfer_amount, trash)
                l += 1
            right_pipette.drop_tip()
        
        #equilibrating with 50 mM HEPES, 50 mM NaCl
        ctx.comment("This is the beginning of equilibrating method.")
        n = 0
        left_pipette.pick_up_tip()
        while n <= len(well_position) - 1:
            left_pipette.flow_rate.aspirate = 200
            left_pipette.flow_rate.dispense = 10
            left_pipette.transfer(
                transfer_amount,
                tube_rack['A2'],
                reaction_plate[f'{well_position[n]}'].bottom(3),
                blow_out=True,
                blowout_location='destination well',
                new_tip='never'
            )
            n += 1
        left_pipette.drop_tip()
        
        ctx.move_labware(labware = reaction_plate,
                                new_location=hs_adapter,
                                use_gripper=True
                                )
        heater_shaker.close_labware_latch()
            
        heater_shaker.set_and_wait_for_shake_speed(shaker)
        ctx.delay(minutes=0.5)
        heater_shaker.deactivate_shaker()    
        heater_shaker.open_labware_latch()
        ctx.move_labware(labware = reaction_plate,
                            new_location=mm_adapter,
                            use_gripper=True
                            )
        ctx.delay(minutes=isolation_time)

            # wash removal using 8 channel pipettes
        l = 1
        right_pipette.pick_up_tip()
        while l <= number_of_lane:
            right_pipette.flow_rate.aspirate = 10
            right_pipette.flow_rate.dispense = 200
            right_pipette.aspirate(transfer_amount, reaction_plate[f'A{l}'])
            right_pipette.dispense(transfer_amount, trash)
            l += 1
        right_pipette.drop_tip()


    def ending():
        ctx.comment("This is the beginning of ending method.")
        n = 0
        left_pipette.pick_up_tip()
        while n <= len(well_position) - 1:
            left_pipette.flow_rate.aspirate = 200
            left_pipette.flow_rate.dispense = 10
            left_pipette.transfer(
                transfer_amount,
                tube_rack['A2'],
                reaction_plate[f'{well_position[n]}'].bottom(3),
                blow_out=True,
                blowout_location='destination well',
                new_tip='never'
            )
            n += 1
        left_pipette.drop_tip()
        heater_shaker.open_labware_latch()
        ctx.comment("This is the end of ending method.")
    
    current_cycle = 1
    number_of_steps = ub_len -  acceptor
    
    ctx.pause('place the reaction plate to magnetic module')

    while current_cycle <= number_of_steps:
        #, PLP solution information
        deprotection(source_column=current_cycle-1) #, PLP solution information
        wash_deprotection(source_column=(number_of_lane))
        reaction(source_column=(number_of_lane, current_cycle))
        wash_reaction(source_column=(number_of_lane, current_cycle))
        current_cycle += 1
    
    ending()
        '''
    
    # Write to BytesIO as a .py file
    opentrons_bytes = io.BytesIO()
    opentrons_bytes.write(opentrons_code.encode('utf-8'))
    opentrons_bytes.seek(0)
    # Set a name attribute for FastAPI StreamingResponse compatibility
    opentrons_bytes.name = 'opentrons.py'
    return opentrons_bytes

def build_reaction_dictionaries_for_UI(data_dict, indexes, multimer_size):

    import re

    def format_text_with_bold(text):
        # Unicode superscript/subscript mapping
        subscript_map = str.maketrans("0123456789", "₀₁₂₃₄₅₆₇₈₉")
        superscript_map = {
            '0': '⁰', '1': '¹', '2': '²', '3': '³', '4': '⁴',
            '5': '⁵', '6': '⁶', '7': '⁷', '8': '⁸', '9': '⁹',
            'a': 'ᵃ', 'b': 'ᵇ', 'c': 'ᶜ', 'd': 'ᵈ', 'e': 'ᵉ',
            'f': 'ᶠ', 'g': 'ᵍ', 'h': 'ʰ', 'i': 'ⁱ', 'j': 'ʲ',
            'k': 'ᵏ', 'l': 'ˡ', 'm': 'ᵐ', 'n': 'ⁿ', 'o': 'ᵒ',
            'p': 'ᵖ', 'r': 'ʳ', 's': 'ˢ', 't': 'ᵗ', 'u': 'ᵘ',
            'v': 'ᵛ', 'w': 'ʷ', 'x': 'ˣ', 'y': 'ʸ', 'z': 'ᶻ',
            'A': 'ᴬ', 'B': 'ᴮ', 'D': 'ᴰ', 'E': 'ᴱ', 'G': 'ᴳ',
            'H': 'ᴴ', 'I': 'ᴵ', 'J': 'ᴶ', 'K': 'ᴷ', 'L': 'ᴸ',
            'M': 'ᴹ', 'N': 'ᴺ', 'O': 'ᴼ', 'P': 'ᴾ', 'R': 'ᴿ',
            'T': 'ᵀ', 'U': 'ᵁ', 'V': 'ⱽ', 'W': 'ᵂ'
        }

        # Convert _<digit> to subscript
        text = re.sub(r'_(\d)', lambda m: m.group(1).translate(subscript_map), text)
        # Convert ^<char> to superscript
        text = re.sub(r'\^([A-Za-z0-9])', lambda m: superscript_map.get(m.group(1), m.group(1)), text)
        # Wrap any number immediately following a superscript with {BOLD} marker
        superscript_chars = ''.join(superscript_map.values())
        # Matches a superscript character from your superscript_chars.
        # Ensures it’s followed by a space and digits.
        # Only wraps the digits (\2) in {BOLD}, not the superscript or the space.    
        text = re.sub(r'([' + re.escape(superscript_chars) + r']) (\d+)', r'\1 {BOLD}\2{BOLD}', text)    
        return text

    def format_text(text):
        # Unicode superscript/subscript mapping
        subscript_map = str.maketrans("0123456789", "₀₁₂₃₄₅₆₇₈₉")
        superscript_map = {
            '0': '⁰', '1': '¹', '2': '²', '3': '³', '4': '⁴',
            '5': '⁵', '6': '⁶', '7': '⁷', '8': '⁸', '9': '⁹',
            'a': 'ᵃ', 'b': 'ᵇ', 'c': 'ᶜ', 'd': 'ᵈ', 'e': 'ᵉ',
            'f': 'ᶠ', 'g': 'ᵍ', 'h': 'ʰ', 'i': 'ⁱ', 'j': 'ʲ',
            'k': 'ᵏ', 'l': 'ˡ', 'm': 'ᵐ', 'n': 'ⁿ', 'o': 'ᵒ',
            'p': 'ᵖ', 'r': 'ʳ', 's': 'ˢ', 't': 'ᵗ', 'u': 'ᵘ',
            'v': 'ᵛ', 'w': 'ʷ', 'x': 'ˣ', 'y': 'ʸ', 'z': 'ᶻ',
            'A': 'ᴬ', 'B': 'ᴮ', 'D': 'ᴰ', 'E': 'ᴱ', 'G': 'ᴳ',
            'H': 'ᴴ', 'I': 'ᴵ', 'J': 'ᴶ', 'K': 'ᴷ', 'L': 'ᴸ',
            'M': 'ᴹ', 'N': 'ᴺ', 'O': 'ᴼ', 'P': 'ᴾ', 'R': 'ᴿ',
            'T': 'ᵀ', 'U': 'ᵁ', 'V': 'ⱽ', 'W': 'ᵂ'
        }

        # Convert _<digit> to subscript
        text = re.sub(r'_(\d)', lambda m: m.group(1).translate(subscript_map), text)
        # Convert ^<char> to superscript
        text = re.sub(r'\^([A-Za-z0-9])', lambda m: superscript_map.get(m.group(1), m.group(1)), text)
        # Wrap any number immediately following a superscript with {BOLD} marker
        superscript_chars = ''.join(superscript_map.values())
        # Matches a superscript character from your superscript_chars.
        # Ensures it’s followed by a space and digits.
        # Only wraps the digits (\2) in {BOLD}, not the superscript or the space.    
        text = re.sub(r'([' + re.escape(superscript_chars) + r']) (\d+)', r'\1 \2', text)    
        return text

    def single_dict_for_reaction_schemes(data_dict, idx, multimer_size):
        
        """ 
        Function to create a dictionary for reaction schemes based on the provided index and multimer size.
        This function retrieves the combined database and context history from the data dictionary,
        filters the combined database for the specified index, and extracts relevant columns to form a dictionary.
        
        Parameters:
        
        - data_dict (dict): Dictionary containing the combined database and context history.
        - idx (int): The index for which the reaction scheme is to be created.  
        - multimer_size (int): The size of the multimer (e.g., 4 for tetramers, 5 for pentamers).
        
        Returns:
        
        - steps_dict (list): A list containing a dictionary with the reaction scheme details.
        The dictionary includes the simulation index, multimer ID, acceptor dimer code, and combined values
        for the steps in the reaction scheme.
        The keys in the dictionary are updated to have more readable names for the UI.
        The values are formatted to include the acceptor dimer code and combined values for each step
        in the reaction scheme.
        The function also handles the representation of different steps in the reaction scheme,
        such as deprotection and formation steps, and updates the values accordingly.   
        
        """

        combined_database = data_dict['combined_database']
        context_history = data_dict['context_history']

        # Get the acceptor dimer name from the context history

        def get_acceptor_dimer_name(context_history, idx):
            """ Get the acceptor dimer code from the working DataFrame. """
            
            # Create a dimer encoded dictionary
            dimer_encoded_str_dictionary = {
                # maybe change the code
                'his-GG-1ubq-1-(<K48_1ubq-2-(<K48_SMAC><K63_ABOC>)>)' : "Ub₂ᴬ 9",
                'his-GG-1ubq-1-(<K48_1ubq-2-(<K48_ABOC><K63_SMAC>)>)' : "Ub₂ᴬ 10",
                'his-GG-1ubq-1-(<K48_1ubq-2-(<K48_ABOC><K63_ABOC>)>)' : "Ub₂ᴬ 11",
                'his-GG-1ubq-1-(<K48_1ubq-2-(<K48_SMAC>)><K63_ABOC>)': "Ub₂ᴬ 12",
                'his-GG-1ubq-1-(<K48_1ubq-2-(<K48_SMAC><K63_ABOC>)><K63_ABOC>)': "Ub₂ᴬ 13",
                'his-GG-1ubq-1-(<K48_1ubq-2-(<K48_ABOC><K63_SMAC>)><K63_ABOC>)': "Ub₂ᴬ 14",
                'his-GG-1ubq-1-(<K63_1ubq-2-(<K48_SMAC><K63_ABOC>)>)': "Ub₂ᴬ 15",
                'his-GG-1ubq-1-(<K63_1ubq-2-(<K48_ABOC><K63_SMAC>)>)': "Ub₂ᴬ 16",
                'his-GG-1ubq-1-(<K63_1ubq-2-(<K48_ABOC><K63_ABOC>)>)': "Ub₂ᴬ 17",
                'his-GG-1ubq-1-(<K48_ABOC><K63_1ubq-2-(<K63_SMAC>)>)': "Ub₂ᴬ 18",
                'his-GG-1ubq-1-(<K48_ABOC><K63_1ubq-2-(<K48_SMAC><K63_ABOC>)>)': "Ub₂ᴬ 19",
                'his-GG-1ubq-1-(<K48_ABOC><K63_1ubq-2-(<K48_ABOC><K63_SMAC>)>)': "Ub₂ᴬ 20",
            }

            working_context_df = context_history[context_history['index'] == idx]

            # Get the value of 'context' in 'dimer_formation' column for the selected index
            if not working_context_df.empty:
                context_dict = working_context_df.iloc[0]['dimer_formation']
                # If the value is a string, try to parse it as a dictionary
                if isinstance(context_dict, str):
                    import ast
                    try:
                        context_dict = ast.literal_eval(context_dict)
                    except Exception:
                        context_dict = {}
                multimer_name = context_dict.get('multimer_string_name', None)
                # Map the multimer name to the dimer code using the dictionary
                dimer_code = dimer_encoded_str_dictionary.get(multimer_name, None)
                return dimer_code
            else:
                ValueError(f"No context row found for index: {idx}")
                return None

        # Get the acceptor dimer name for the given index
        multimer_code = get_acceptor_dimer_name(context_history, idx)

        # ============================================================

        # Filter combined_database for rows where index == idx and table_origin is 'Donors' or 'Reactions'
        filtered_combined_df = combined_database[(combined_database['index'] == idx) & (combined_database['table_origin'].isin(['Donors', 'Reactions']))]

        # Determine the last column name based on multimer size
        if multimer_size == 4:
            last_col = 'tetramer_formation'
        elif multimer_size == 5:
            last_col = 'pentamer_formation'
        else:
            raise ValueError(f"Unsupported multimer_size: {multimer_size}")

        # Get all column names
        all_cols = list(filtered_combined_df.columns)

        # Find the start and end indices for slicing columns
        try:
            start_idx = all_cols.index('dimer_deprotection')
            end_idx = all_cols.index(last_col)
        except ValueError as e:
            raise ValueError(f"Column not found: {e}")

        # Slice the columns between dimer_deprotection and last_col (inclusive)
        selected_cols = all_cols[start_idx:end_idx+1]

        # Get the relevant part of the DataFrame
        selected_steps_df = filtered_combined_df[selected_cols]

        # Combine Donors and Reactions rows to form combined_values like "Ub^D 2\ngp78-Ube2g2"
        combined_values = []
        if not selected_steps_df.empty and len(selected_steps_df) == 2:
            # Assume first row is Donors, second is Reactions (order from filter)
            donors_row = selected_steps_df.iloc[0]
            reactions_row = selected_steps_df.iloc[1]
            for donor_val, reaction_val in zip(donors_row, reactions_row):
                combined = f"{donor_val}\n{reaction_val}"
                combined_values.append(combined)
        else:
            ValueError("selected_steps_df does not have exactly 2 rows (Donors and Reactions). Cannot combine.")

        # This is just syntax for the UI
        representation_changes = {
                'Ubc13/Mms2' : 'Ubc13-Mms2',
                'gp78/Ube2g2' : 'gp78-Ube2g2',
                'ubi_ubq_1_K48_ABOC_K63_ABOC' : 'Ubᴰ 5',
                'ubi_ubq_1_K48_SMAC_K63_ABOC' : 'Ubᴰ 2',
                'ubi_ubq_1_K48_ABOC_K63_SMAC' : 'Ubᴰ 4',
                'ubi_ubq_1_K48_SMAC' : 'Ubᴰ 1',
                'ubi_ubq_1_K63_SMAC' : 'Ubᴰ 3',
                'nan\nSMAC_deprot' : 'Smac',
                'nan\nFAKE_deprot' : 'Fake', 
                'dimer_deprotection' : 'Dimer\ndeprotection',
                'trimer_deprotection' : 'Trimer\ndeprotection',
                'tetramer_deprotection' : 'Tetramer\ndeprotection',
                'pentamer_deprotection' : 'Pentamer\ndeprotection',
                'trimer_formation' : 'Trimer\nformation',
                'tetramer_formation' : 'Tetramer\nformation',
                'pentamer_formation' : 'Pentamer\nformation',
            }

        def update_values(values, representation_changes):
            """
            Updates a list of values by replacing exact matches or substrings
            based on a given mapping.

            Parameters:
            - values (list of str): The original values.
            - representation_changes (dict): Mapping of old -> new representations.

            Returns:
            - list of str: Updated values.
            """
            updated_values = []

            for val in values:
                if val in representation_changes:
                    updated_val = representation_changes[val]
                else:
                    updated_val = val
                    for old, new in representation_changes.items():
                        if old in updated_val:
                            updated_val = updated_val.replace(old, new)
                updated_values.append(updated_val)

            return updated_values


        updated_values = update_values(combined_values, representation_changes)
        headers_for_dict = update_values(selected_steps_df.columns, representation_changes)


        # Get the unique multimer_id value from filtered_combined_df
        multimer_id = filtered_combined_df['multimer_id'].unique()[0]

        def format_multimer_id(multimer_id):
            """
            Reformat multimer_id string from 'UbX_Y' to 'Ub_X Y', with suffixes for certain sizes.
            For size 4: add ^T as superscript. For size 5: add ^P as superscript.
            """
            match = re.match(r'^Ub(\d+)_(\d+)$', multimer_id)
            if match:
                size, number = match.group(1), match.group(2)
                suffix = '^T' if size == '4' else '^P' if size == '5' else ''
                return f"Ub_{size}{suffix} {number}"
            return multimer_id
        multimer_id = format_multimer_id(multimer_id)
        multimer_id = format_text(multimer_id)

        # Create a dictionary mapping columns to updated values, with 'Acceptor' as the first key
        steps_dict = {
            'Simulation\nindex': idx, 
            'Multimer Id': multimer_id,
            'Acceptor': multimer_code
            }

        # Update the steps_dict with the combined values
        steps_dict.update(dict(zip(headers_for_dict, updated_values)))
        
        # List the steps_dict to be returned
        steps_dict = [steps_dict]
        # Ensure 'Reaction Number' is always the first key in each dictionary before converting to DataFrame
        steps_dict[0] = {"Reaction\nNumber": None, **steps_dict[0]}

        return steps_dict

    def full_dict_df_for_reaction_schemes(data_dict, indexes, multimer_size):
        """
        Function to create a dictionary for reaction schemes based on the provided index and multimer size.
        This function retrieves the combined database and context history from the data dictionary,
        filters the combined_database for the specified index, and extracts relevant columns to form a dictionary.
        """

        steps_full_dict = []
        for i in range(len(indexes)):
            idx = indexes[i]
            # Call the placeholder function to get the steps_dict for each index
            # This is just a placeholder function to avoid syntax errors
            # In practice, this would be replaced with the actual function that processes the data
            steps_dict = single_dict_for_reaction_schemes(data_dict, idx, multimer_size)
            steps_full_dict += steps_dict

        # Add reaction numbers to each step in steps_full_dict
        for i, step in enumerate(steps_full_dict, start=1):
            step['Reaction\nNumber'] = str(i)

        # Create DataFrame preserving original column names
        steps_df = pd.DataFrame(steps_full_dict)

        # Insert headers as the first row
        header_row = {col: col for col in steps_df.columns}
        steps_df = pd.concat([pd.DataFrame([header_row]), steps_df], ignore_index=True)

        return steps_df, steps_full_dict

    def generate_all_lines_dicts(data_dict, indexes, multimer_size):
        all_lines_dicts = []
        ubiquitin_history = data_dict['ubiquitin_history']
        donor_history = data_dict['donor_history']

        for idx in indexes:
            line = ubiquitin_history[ubiquitin_history['index'] == idx]
            line_dicts = line.to_dict(orient='records')
            donor_line = donor_history[donor_history['index'] == idx]
            donor_line_dicts = donor_line.to_dict(orient='records')

            for line_dict in line_dicts:
                steps_dict = next((step for step in steps_full_dict if step['Simulation\nindex'] == idx), {})
                line_dict = {f"ubi_his_JSON_{k}": v for k, v in line_dict.items()}
                overlapping_keys = set(line_dict.keys()) & set(steps_dict.keys())
                if overlapping_keys:
                    print(f"Overlapping keys for index {idx}: {overlapping_keys}")
                line_dict.update(steps_dict)

                for donor_line_dict in donor_line_dicts:
                    donor_line_dict = {f"donor_JSON_{k}": v for k, v in donor_line_dict.items() if 'formation' in k}
                    line_dict.update(donor_line_dict)

                all_lines_dicts.append(line_dict)

        return all_lines_dicts

    def compress_duplicates(all_lines_dicts):
        compressed_dicts = []
        seen_items = {}

        for index, item in enumerate(all_lines_dicts):
            item_str = json.dumps(item, sort_keys=True)
            if item_str in seen_items:
                seen_items[item_str]['end'] = index + 1
            else:
                seen_items[item_str] = {'start': index + 1, 'end': index + 1}
                compressed_dicts.append(item)

        for item in compressed_dicts:
            item_str = json.dumps(item, sort_keys=True)
            range_info = seen_items[item_str]
            if range_info['start'] != range_info['end']:
                item['Reaction\nNumber'] = f"{range_info['start']} - {range_info['end']}"

        return compressed_dicts

    # Create the full dictionary and DataFrame for reaction schemes
    steps_df, steps_full_dict = full_dict_df_for_reaction_schemes(data_dict, indexes, multimer_size)

    # Call the function
    reaction_schemes_dicts = generate_all_lines_dicts(data_dict, indexes, multimer_size)

    # Compress duplicates in all_lines_dicts
    reaction_schemes_dicts = compress_duplicates(reaction_schemes_dicts)

    return reaction_schemes_dicts


def ubiquitin_building_wo_iterate(
    parent_dictionary: dict | str,
    ubi_molecule_to_add: dict | str,
    bonding_ubiquitin_number: int,
    new_ubiquitin_number: int,
    lysine_residue: str
) -> dict:
    """
    Entry point for building a ubiquitin chain by attaching a molecule or protecting group.

    Args:
        parent_dictionary (dict or str): Input protein structure.
        ubi_molecule_to_add (dict or str): Ubiquitin or protecting group (SMAC/ABOC).
        bonding_ubiquitin_number (int): The chain number to which the molecule is added.
        lysine_residue (str): The specific lysine site for conjugation.

    Returns:
        dict: The updated protein dictionary with changes applied.
    """

    import src.main as main
    import src.utils.utils as utils
    import src.utils.logging_utils as logging_utils
    from src.utils.utils import convert_json_to_dict
    from src.utils.logging_utils import log_end_of_branching, log_end_of_protein, log_protein_details, log_branching_details
    import copy
    import logging

    # Initialize context for chain numbering and length tracking
    context = {
        "chain_number_list": [1],
        "chain_length_list": [],
    }

    # Normalize input to dicts
    parent_dictionary = convert_json_to_dict(parent_dictionary)

    # Only convert to dict if not a protecting group
    if ubi_molecule_to_add not in ('SMAC', 'ABOC'):
        ubi_molecule_to_add = convert_json_to_dict(ubi_molecule_to_add)

    new_ubi_molecule_to_add = copy.deepcopy(ubi_molecule_to_add)
    new_ubi_molecule_to_add['chain_number'] = new_ubiquitin_number

     # Error handling fop invalid ubi_molecule_to_add
    if ubi_molecule_to_add not in ('SMAC', 'ABOC'):
        if not isinstance(ubi_molecule_to_add, dict):
            raise TypeError("ubi_molecule_to_add must be a dictionary or 'SMAC'/'ABOC' string")

    def inner_wrapper_ubiquitin_building_wo_iterate(
        input_dictionary: dict | str,
        ubi_molecule_to_add: dict | str,
        bonding_ubiquitin_number: int,
        new_ubiquitin_number: int,
        lysine_residue: str,
        context: dict
    ) -> tuple:
        """
        Recursively traverse and modify the ubiquitin structure to add new branches
        or protecting groups to a given lysine site.

        Args:
            input_dictionary (dict or str): The protein or ubiquitin structure.
            ubi_molecule_to_add (dict or str): Ubiquitin or protecting group.
            bonding_ubiquitin_number (int): The specific chain number to modify.
            lysine_residue (str): Lysine site to target.
            context (dict): Tracks chain numbering and lengths.

        Returns:
            tuple: Updated working dictionary and context.
        """
        # Deep copy to avoid mutating input
        working_dictionary = copy.deepcopy(input_dictionary)
        working_dictionary = convert_json_to_dict(working_dictionary)

        # Set the current chain number from context
        # working_dictionary['chain_number'] = context['chain_number_list'][-1]
        # Set and record chain length
        working_dictionary['chain_length'] = len(working_dictionary['FASTA_sequence'])
        context['chain_length_list'].append(working_dictionary['chain_length'])

        # Log protein details for debugging and traceability
        log_protein_details(working_dictionary, context)

        # Increment chain_number for future recursive calls
        context['chain_number_list'].append(context['chain_number_list'][-1] + 1)

        for bra in working_dictionary['branching_sites']:
            log_branching_details(bra, working_dictionary, context)

            # Apply modification logic to the lysine site
            bra, working_dictionary = main.handle_lysine_modification(
                bra, working_dictionary, ubi_molecule_to_add, bonding_ubiquitin_number, lysine_residue
            )

            # Log the state of the current branch
            if bra['children'] in ('SMAC', 'ABOC'):
                logging.info(f"Protecting Group: {bra['children']}")
            elif bra['children'] == "":
                logging.info(f"There is no Protecting Group on: {bra['site_name']}")
            elif isinstance(bra['children'], dict):
                logging.info(f"NEXT CHAIN: {bra['children']}")
                # Recursively process the next ubiquitin chain
                bra['children'], context = inner_wrapper_ubiquitin_building_wo_iterate(
                    bra['children'], ubi_molecule_to_add, bonding_ubiquitin_number, new_ubiquitin_number, lysine_residue, context
                )
            log_end_of_branching()

        log_end_of_protein(working_dictionary)

        return working_dictionary, context

    # Build structure recursively
    output_dictionary, output_context = inner_wrapper_ubiquitin_building_wo_iterate(
        parent_dictionary, new_ubi_molecule_to_add, bonding_ubiquitin_number, new_ubiquitin_number, lysine_residue, context
    )
    
    return output_dictionary, output_context

def get_indexes_for_final_multimer(json_output, ubiquitin_history):
    """
    Match the final multimer from JSON output to the ubiquitin history and return unique indices.
    Args:
        json_output (list): List of dictionaries containing 'from', 'to', and 'linkage'.
        ubiquitin_history (DataFrame): DataFrame containing the history of ubiquitin structures.    
    Returns:
        list: Unique indices corresponding to the final multimer.
    """
    
    import src.main as main

    ubiquitin_monomer = {
        "protein": "1ubq",
        "chain_number": 1,
        "FASTA_sequence": "MQIFVKTLTGKTITLEVEPSDTIENVKAKIQDKEGIPPDQQRLIFAGKQLEDGRTLSDYNIQKESTLHLVLRLRGG",
        "chain_length": 76,
        "branching_sites": [
            {"site_name": "M1", "sequence_id": "(M)QIF", "children": ""},
            {"site_name": "K6", "sequence_id": "IFV(K)TLT", "children": ""},
            {"site_name": "K11", "sequence_id": "LTG(K)TIT", "children": ""},
            {"site_name": "K27", "sequence_id": "ENV(K)AKI", "children": ""},
            {"site_name": "K29", "sequence_id": "VKA(K)IQD", "children": ""},
            {"site_name": "K33", "sequence_id": "IQD(K)EGI", "children": ""},
            {"site_name": "K48", "sequence_id": "FAG(K)QLE", "children": ""},
            {"site_name": "K63", "sequence_id": "NIQ(K)EST", "children": ""}
        ]
    }

    histag_ubiquitin_monomer = {
        "protein": "1ubq",
        "chain_number": 0,
        "FASTA_sequence": "MQIFVKTLTGKTITLEVEPSDTIENVKAKIQDKEGIPPDQQRLIFAGKQLEDGRTLSDYNIQKESTLHLVLRLRGGDHHHHHH",
        "chain_length": 83,
        "branching_sites": [
            {"site_name": "M1", "sequence_id": "(M)QIF", "children": ""},
            {"site_name": "K6", "sequence_id": "IFV(K)TLT", "children": ""},
            {"site_name": "K11", "sequence_id": "LTG(K)TIT", "children": ""},
            {"site_name": "K27", "sequence_id": "ENV(K)AKI", "children": ""},
            {"site_name": "K29", "sequence_id": "VKA(K)IQD", "children": ""},
            {"site_name": "K33", "sequence_id": "IQD(K)EGI", "children": ""},
            {"site_name": "K48", "sequence_id": "FAG(K)QLE", "children": ""},
            {"site_name": "K63", "sequence_id": "NIQ(K)EST", "children": ""}
        ]
    }

    for index, json_dict in enumerate(json_output):
        if index == 0:
            growing_ubiquitin, growing_context = ubiquitin_building_wo_iterate(
                histag_ubiquitin_monomer,
                ubiquitin_monomer,
                json_dict['from'],
                json_dict['to'],
                json_dict['linkage']
            )
        else:
            growing_ubiquitin, growing_context = ubiquitin_building_wo_iterate(
                growing_ubiquitin,
                ubiquitin_monomer,
                json_dict['from'],
                json_dict['to'],
                json_dict['linkage']
            )

    final_ubiquitin, final_context = main.iterate_through_ubiquitin(growing_ubiquitin)
    indexes = ubiquitin_history.loc[ubiquitin_history["final_multimer"] == str(final_ubiquitin), "index"].dropna().unique()
    return [int(i) for i in indexes]
