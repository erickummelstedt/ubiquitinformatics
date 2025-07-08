import json 
import logging
import copy
import sys
import ast
import numpy as np
from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import font_manager
import ast 
import subprocess

# Dynamically get the backend path relative to this script location
script_path = Path(__file__).resolve()
project_root = script_path.parents[2]  # Go up to project root (adjust if needed)
src_path = project_root / 'back_end' / 'src'
sys.path.insert(0, str(src_path))

from utils.utils import *
from utils.logging_utils import *
from main import *
from plotting import *

multimer_size = 4

# ≠===========================================================
# Functions
# ≠===========================================================

# Create plate dataframes for the selected indexes
def count_nonzero_values(df):
    """
    Counts occurrences of non-zero values in a DataFrame.

    Parameters:
        df (pd.DataFrame): The input DataFrame.

    Returns:
        dict: Dictionary of value counts, excluding zeros.
    """
    counts = pd.Series(df.values.ravel())
    return counts[counts != 0].value_counts().sort_index().to_dict()

# download CSV files
def download_data_dict(multimer_size):
    input_dir = project_root / 'back_end' / 'data' / 'filtered_reaction_database' / f'multimer_size_{multimer_size}'

    combined_database = pd.read_csv(input_dir / 'combined_database.csv', index_col=0)
    context_history = pd.read_csv(input_dir / 'context_history.csv', index_col=0)
    donor_history = pd.read_csv(input_dir / 'donor_history.csv', index_col=0)
    reaction_history = pd.read_csv(input_dir / 'reaction_history.csv', index_col=0)
    ubiquitin_history = pd.read_csv(input_dir / 'ubiquitin_history.csv', index_col=0)

    return {
        'combined_database': combined_database,
        'context_history': context_history,
        'donor_history': donor_history,
        'reaction_history': reaction_history,
        'ubiquitin_history': ubiquitin_history
    }

# Create plate dataframes for the selected indexes
data_dict = download_data_dict(multimer_size)
combined_database = data_dict['combined_database']
context_history = data_dict['context_history']
donor_history = data_dict['donor_history']
reaction_history = data_dict['reaction_history']
ubiquitin_history = data_dict['ubiquitin_history']

# Create the plate dataframes for the selected indexes
# Select the indexes of the multimer_ids that are used in synthesis
selected_ids = ["Ub4_5","Ub4_10","Ub4_2","Ub4_2","Ub4_2","Ub4_2","Ub4_2","Ub4_2","Ub4_2","Ub4_2","Ub4_8", "Ub4_11","Ub4_7"]
indexes = list()

# Get the indexes of the selected multimer_ids that are used in synthesis
for id in selected_ids:
    new_index = int(combined_database[(combined_database['multimer_id'] == id) & (combined_database['used_in_synthesis'] == 1)]['index'].unique()[0])
    indexes.append(new_index)

# Create the plate dataframes for the selected indexes
output_dict = inner_create_plate_dfs(data_dict, indexes, multimer_size)

# Extract the plate dataframes from the output dictionary
enzymes_donors_96 = output_dict['enzymes_donors_96']
deprots_96 = output_dict['deprots_96']
acceptors_96 = output_dict['dimer_acceptors_96']

# pull dictionaries from the output_dict
e_d_encoded_dictionary = output_dict['e_d_encoded_dictionary']
deprot_encoded_dictionary = output_dict['deprot_encoded_dictionary']
dimer_encoded_dictionary = output_dict['dimer_encoded_dictionary']

# Set up is done, now we can count the non-zero values in each DataFrame
# ============================================================

# Count non-zero values in each DataFrame
enzymes_donors_96_counts = count_nonzero_values(enzymes_donors_96)
deprots_96_counts = count_nonzero_values(deprots_96)
acceptors_96_counts = count_nonzero_values(acceptors_96)



from collections import defaultdict

def count_mixture_and_components(e_d_encoded_dict, count_dict, component_list):
    # Count per mixture (decoded from keys)
    mixture_counts = defaultdict(int)
    component_counts = defaultdict(int)

    # Reverse lookup: map encoded values back to mixture keys
    reversed_dict = {v: k for k, v in e_d_encoded_dict.items()}

    # Ensure every mixture from the dictionary is present, even if not in count_dict
    for encoded_key, mixture in reversed_dict.items():
        mixture_counts[mixture] = 0  # Initialize all mixtures to 0

    for encoded_key, count in count_dict.items():
        if encoded_key not in reversed_dict:
            continue  # skip unknown keys
        mixture = reversed_dict[encoded_key]
        mixture_counts[mixture] += count
        # Flatten the components in the mixture (tuple of (substrate, enzyme))
        for item in mixture:
            for component in component_list:
                if component == item or (isinstance(item, str) and component in item):
                    component_counts[component] += count

    # Also return the mapping from mixture (tuple) to encoded code
    mixture_code_map = {mixture: code for code, mixture in e_d_encoded_dict.items()}
    return dict(mixture_counts), dict(component_counts), mixture_code_map

components = [
    'Ubc13/Mms2',
    'gp78/Ube2g2',
    'Ube2K',
    'ubi_ubq_1_K48_ABOC_K63_ABOC',
    'ubi_ubq_1_K48_SMAC_K63_ABOC',
    'ubi_ubq_1_K48_ABOC_K63_SMAC',
    'ubi_ubq_1_K48_SMAC',
    'ubi_ubq_1_K63_SMAC'
]

default_final_concentrations = {
    # E1 enzyme
    'hUba1': 1,  # in µM
    # E2 enzymes
    'Ubc13/Mms2': 20,  # in µM
    'gp78/Ube2g2': 20,  # in µM
    'Ube2K': 20,  # in µM
    'ubi_ubq_1_K48_ABOC_K63_ABOC': 20, # in µM
    'ubi_ubq_1_K48_SMAC_K63_ABOC': 20, # in µM
    'ubi_ubq_1_K48_ABOC_K63_SMAC': 20, # in µM
    'ubi_ubq_1_K48_SMAC': 20, # in µM
    'ubi_ubq_1_K63_SMAC': 20, # in µM
    # Acceptors
    'Acceptor Dimer Concentration': 10,  # in nmol
    'Reaction Volume [uL]': 200,  # in µL
    # TCEP & ATP
    'TCEP': 1,  # in mM 
    'ATP/Mg2+': 10,  # in mM
}

initial_concentraions = {
    'TCEP': 10,  # in mM 
    'ATP/Mg2+': 100,  # in mM
    'hUba1': 10,  # in µM
}

masses = {
    'hUba1': 116300,  # in Da
    'Ubc13/Mms2': 33800,  # in Da
    'gp78/Ube2g2': 27600,  # in Da
    'Ube2K': 22400,  # in Da
    # Acceptor monomers
    'histag_ubi_ubq_1' : 9527, # in Da
    'histag_ubi_ubq_1_K48_aboc' : 9658, # in Da
    'histag_ubi_ubq_1_K63_aboc' : 9658, # in Da
    # Donors 
    'ubi_ubq_1_K48_ABOC_K63_ABOC': 8851, # in Da 
    'ubi_ubq_1_K48_SMAC_K63_ABOC': 8864, # in Da
    'ubi_ubq_1_K48_ABOC_K63_SMAC': 8864, # in Da
    'ubi_ubq_1_K48_SMAC': 8733.00, # in Da
    'ubi_ubq_1_K63_SMAC': 8733.00, # in Da
}

# Define acceptor description dictionary
acceptor_description_dict = [
    {"No.": "Ub₂ᴬ 9", "Dimer Linkage": "K48", "Proximal Ub K48": "-", "Proximal Ub K63": "K", "Distal Ub K48": "Smac", "Distal Ub K63": "Aboc", "MW (Da)": 18372},
    {"No.": "Ub₂ᴬ 10", "Dimer Linkage": "K48", "Proximal Ub K48": "-", "Proximal Ub K63": "K", "Distal Ub K48": "Aboc", "Distal Ub K63": "Smac", "MW (Da)": 18372},
    {"No.": "Ub₂ᴬ 11", "Dimer Linkage": "K48", "Proximal Ub K48": "-", "Proximal Ub K63": "K", "Distal Ub K48": "Aboc", "Distal Ub K63": "-", "MW (Da)": 18360},
    {"No.": "Ub₂ᴬ 12", "Dimer Linkage": "K48", "Proximal Ub K48": "-", "Proximal Ub K63": "Aboc", "Distal Ub K48": "Smac", "Distal Ub K63": "K", "MW (Da)": 18372},
    {"No.": "Ub₂ᴬ 13", "Dimer Linkage": "K48", "Proximal Ub K48": "-", "Proximal Ub K63": "Aboc", "Distal Ub K48": "Smac", "Distal Ub K63": "-", "MW (Da)": 18504},
    {"No.": "Ub₂ᴬ 14", "Dimer Linkage": "K48", "Proximal Ub K48": "-", "Proximal Ub K63": "Aboc", "Distal Ub K48": "Aboc", "Distal Ub K63": "-", "MW (Da)": 18504},
    {"No.": "Ub₂ᴬ 15", "Dimer Linkage": "K63", "Proximal Ub K48": "K", "Proximal Ub K63": "-", "Distal Ub K48": "Smac", "Distal Ub K63": "Aboc", "MW (Da)": 18372},
    {"No.": "Ub₂ᴬ 16", "Dimer Linkage": "K63", "Proximal Ub K48": "K", "Proximal Ub K63": "-", "Distal Ub K48": "Aboc", "Distal Ub K63": "Smac", "MW (Da)": 18372},
    {"No.": "Ub₂ᴬ 17", "Dimer Linkage": "K63", "Proximal Ub K48": "K", "Proximal Ub K63": "-", "Distal Ub K48": "Aboc", "Distal Ub K63": "-", "MW (Da)": 18360},
    {"No.": "Ub₂ᴬ 18", "Dimer Linkage": "K63", "Proximal Ub K48": "Aboc", "Proximal Ub K63": "-", "Distal Ub K48": "K", "Distal Ub K63": "-", "MW (Da)": 18372},
    {"No.": "Ub₂ᴬ 19", "Dimer Linkage": "K63", "Proximal Ub K48": "Aboc", "Proximal Ub K63": "-", "Distal Ub K48": "Smac", "Distal Ub K63": "-", "MW (Da)": 18504},
    {"No.": "Ub₂ᴬ 20", "Dimer Linkage": "K63", "Proximal Ub K48": "Aboc", "Proximal Ub K63": "-", "Distal Ub K48": "Aboc", "Distal Ub K63": "Smac", "MW (Da)": 18504},
]

# prep_headers change depending on the regent
prep_header_dict = {
    'Ubc13/Mms2' : 'Enzyme',
    'gp78/Ube2g2' : 'Enzyme',
    'Ube2K' : 'Enzyme',
    'ubi_ubq_1_K48_ABOC_K63_ABOC' : 'Donor: Ubᴰ 5',
    'ubi_ubq_1_K48_SMAC_K63_ABOC' : 'Donor: Ubᴰ 2',
    'ubi_ubq_1_K48_ABOC_K63_SMAC' : 'Donor: Ubᴰ 4',
    'ubi_ubq_1_K48_SMAC' : 'Donor: Ubᴰ 1',
    'ubi_ubq_1_K63_SMAC' : 'Donor: Ubᴰ 3',
}

mixture_counts, component_counts, mixture_code_map = count_mixture_and_components(
    e_d_encoded_dictionary, enzymes_donors_96_counts, components)


# Create Excel output file combining mixture_counts and component_counts
def create_combined_xlsx_bytes(mixture_counts, component_counts, filename, all_components=None, e_d_encoded_dictionary=None):
    
    """
    Create an Excel file with two sheets: one for mixture_counts and one for component_counts.
    Ensures all components from all_components are present in the component sheet (with 0 if missing).
    Includes mixture codes in the mixture sheet if provided.
    Adds a sum row at the bottom of the mixture counts.
    """
    
    # Mixed Reagents DataFrame
    mixture_data = []
    for index, (mixture, count) in enumerate(mixture_counts.items()):
        code = e_d_encoded_dictionary.get(mixture)  # Get the value from the encoding dictionary
        reaction_volume = default_final_concentrations.get('Reaction Volume [uL]', 200)
        mixture_data.append({'Donor + Enzyme Mix': mixture, 'Donor + Enzyme Id': code, 'Count': count, 'Reaction Volume (uL)': reaction_volume})
    mixture_df = pd.DataFrame(mixture_data, columns=['Donor + Enzyme Mix', 'Donor + Enzyme Id', 'Count', 'Reaction Volume (uL)'])
    # Add sum row
    sum_row = {'Donor + Enzyme Mix': 'TOTAL', 'Donor + Enzyme Id': '', 'Count': mixture_df['Count'].sum(), 'Reaction Volume (uL)': ''}
    mixture_df = pd.concat([mixture_df, pd.DataFrame([sum_row])], ignore_index=True)
    # Add Excel formula for 'Donor + Enzyme Volume (uL)' column
    from openpyxl import load_workbook
    
    # Start column index for the mixture sheet
    start_column_index = 1  # Start from the second column (B) (0-indexed, so 1 is column B)
    start_row_index = 1  # Start from the second row (0-indexed, so 1 is row 2)
    start_row = start_row_index + 1

    from io import BytesIO
    output = BytesIO()
    # Create the Excel file with pandas
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        mixture_df.to_excel(writer, sheet_name='Mixed Reagents Counts', index=False, startcol=start_column_index, startrow=start_row_index)
        # Write the component sheet as before
        if all_components is not None:
            comp_data = [(comp, component_counts.get(comp, 0)) for comp in all_components]
            component_df = pd.DataFrame(comp_data, columns=['Base Reagent', 'Count'])
        else:
            component_df = pd.DataFrame(list(component_counts.items()), columns=['Base Reagent', 'Count'])
        component_df.to_excel(writer, sheet_name='Base Reagent Counts', index=False, startcol=start_column_index, startrow=start_row_index)

    # Now load the workbook from the BytesIO buffer
    output.seek(0)
    wb = load_workbook(output)
    ws = wb['Mixed Reagents Counts']

    # Insert new column for 'Donor + Enzyme Volume (uL)\n(Count * Reaction Volume (uL) * 1.1)' after 'Reaction Volume (uL)'
    insert_at = len(mixture_df.columns) + start_column_index + 1 # After 4th column (A-D: Mix, Id, Count, Reaction Vol)
    ws.insert_cols(insert_at)
    ws.cell(row=start_row, column=insert_at).value = 'Donor + Enzyme Volume (uL)\n(Count * Reaction Volume (uL) * 1.1)'
    
    # Get 'Reaction Volume (uL)' and 'Count' column indexes
    reaction_vol_col_idx = start_column_index + 4  # Assuming 'Reaction Volume (uL)' is the 4th column (D)   
    count_col_idx = start_column_index + 3  # Assuming 'Count' is the 3rd column (C) 

    # Add formula for each row except the sum row
    for row in range(start_row + 1, ws.max_row):
        count_cell = ws.cell(row=row, column=reaction_vol_col_idx).coordinate  # 'Count' column
        reaction_vol_cell = ws.cell(row=row, column=count_col_idx).coordinate  # 'Reaction Volume (uL)' column
        formula = f"={count_cell}*{reaction_vol_cell}*1.1"
        ws.cell(row=row, column=insert_at).value = formula
    # --- Add new columns for Donor, Donor Volume (uL), Enzyme, Enzyme Volume (uL) as Excel formulas ---
    # Insert Donor and Enzyme columns after 'Donor + Enzyme Volume (uL)\n(Count * Reaction Volume (uL) * 1.1)'
    donor_col = insert_at + 1
    donor_id_col = donor_col + 1
    donor_vol_col = donor_id_col + 1
    enzyme_col = donor_vol_col + 1
    enzyme_vol_col = enzyme_col + 1

    # Insert columns for Donor, Donor Id, Donor Volume, Enzyme, Enzyme Volume
    ws.insert_cols(donor_col, amount=enzyme_vol_col-insert_at)  # Insert 5 columns for Donor, Donor Id, Donor Volume, Enzyme, Enzyme Volume
    ws.cell(row=start_row, column=donor_col).value = 'Donor'    
    ws.cell(row=start_row, column=donor_id_col).value = 'Donor Id'
    ws.cell(row=start_row, column=donor_vol_col).value = 'Donor Volume (uL)'
    ws.cell(row=start_row, column=enzyme_col).value = 'Enzyme'
    ws.cell(row=start_row, column=enzyme_vol_col).value = 'Enzyme Volume (uL)'

    # Fill Donor and Enzyme columns by parsing the tuple string
    import ast
    for row in range(start_row + 1, ws.max_row):
        mix_val = ws.cell(row=row, column=start_column_index+1).value
        try:
            t = ast.literal_eval(mix_val) if isinstance(mix_val, str) and mix_val.startswith('(') else ('','')
            donor, enzyme = t if isinstance(t, tuple) and len(t) == 2 else ('','')
        except Exception:
            donor, enzyme = '', ''
        ws.cell(row=row, column=donor_col).value = donor
        # Set the donor ID based on the donor name
        ws.cell(row=row, column=donor_id_col).value = prep_header_dict.get(donor)
        ws.cell(row=row, column=enzyme_col).value = enzyme
        # Set Excel formulas for volumes: (Count * Reaction Volume (uL) * 1.1) / 2
        count_cell = ws.cell(row=row, column=count_col_idx).coordinate
        reaction_vol_cell = ws.cell(row=row, column=reaction_vol_col_idx).coordinate
        # Set formulas for Donor Volume and Enzyme Volume
        ws.cell(row=row, column=donor_vol_col).value = f"=({count_cell}*{reaction_vol_cell}*1.1)/2"
        ws.cell(row=row, column=enzyme_vol_col).value = f"=({count_cell}*{reaction_vol_cell}*1.1)/2"
    
    # Update header formatting and auto-fit for all columns (including new ones)
    from openpyxl.styles import Font, Border, Side, Alignment
    header_font = Font(bold=True)
    border = Border(bottom=Side(border_style="medium"))#, left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"))
    for col in range(start_column_index + 1, ws.max_column + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    # Auto-fit all columns to the width of their titles or widest value, accounting for wrapping
    for index, col in enumerate(ws.columns):
        if index < start_column_index:
            continue # Skip the columns before the start_column_index
        header_cell = col[start_column_index]  # Header cell is the first row in the column
        col_letter = header_cell.column_letter
        if header_cell.value == 'Mass Required (mg)\n(MW (Da) * Amount Needed (nmol) / 1000)':
            ws.column_dimensions[col_letter].width = 10  # Set a tighter width for this column
            continue
        if '\n' in str(header_cell.value):
            lines = str(header_cell.value).split('\n')
            max_length = max(len(line) for line in lines)
        else:
            max_length = len(str(header_cell.value)) if header_cell.value is not None else 0
        for cell in col[start_column_index+1:]:
            cell_length = len(str(cell.value)) if cell.value is not None else 0
            if cell_length > max_length:
                max_length = cell_length
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width
    # Highlight non-zero rows in specified columns with light green fill (excluding 'Total' row and 'Donor + Enzyme Mix')
    from openpyxl.styles import PatternFill
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    # Get column indices for the specified columns (excluding 'Donor + Enzyme Mix')
    col_names = [
        'Donor + Enzyme Id', 'Donor', 'Donor Id','Donor Volume (uL)', 'Enzyme', 'Enzyme Volume (uL)'
    ]
    col_indices = {}
    for col in range(start_column_index + 1, ws.max_column + 1):
        header = ws.cell(row=start_row, column=col).value
        if header in col_names:
            col_indices[header] = col
    # Highlight rows where 'Count' (col 3) is nonzero and not the sum row
    for row in range(start_row+1, ws.max_row+1):
        count_val = ws.cell(row=row, column=count_col_idx).value
        first_col_val = ws.cell(row=row, column=start_column_index+1).value
        try:
            is_nonzero = float(count_val) != 0
        except Exception:
            is_nonzero = False
        is_total = str(first_col_val).strip().upper() == 'TOTAL'
        if is_nonzero and not is_total:
            for colname in col_names:
                col = col_indices.get(colname)
                if col:
                    cell = ws.cell(row=row, column=col)
                    cell.fill = green_fill
                    cell.font = Font(bold=True)  # Make highlighted cell bold
    
    count_idx = 2
    donor_idx = 3
    reaction_vol_idx = donor_idx + 1
    total_vol_idx = reaction_vol_idx + 1
    final_conc_idx = total_vol_idx + 1
    amount_needed_idx = final_conc_idx + 1
    mw_idx = amount_needed_idx + 1
    mass_required_idx = mw_idx + 1

    # Format Base Reagent Counts sheet
    ws_comp = wb['Base Reagent Counts']

    # Insert 'Donor ID' column after Counts
    ws_comp.insert_cols(start_column_index + donor_idx)  # Insert after 'Count' (which is now at col 2)
    ws_comp.cell(row=start_row_index + 1, column=start_column_index + donor_idx).value = 'Donor Id'
    # Insert 'Reaction Volume (uL)' column after 'Donor Id'
    ws_comp.insert_cols(start_column_index + reaction_vol_idx)  # Insert after 'Count' (which is now at col 2)
    ws_comp.cell(row=start_row_index + 1, column=start_column_index + reaction_vol_idx).value = 'Reaction Volume (uL)'
    # Insert 'Total Volume (uL)\n(Count * Reaction Volume (uL) * 1.1 / 2)' after 'Reaction Volume (uL)'
    ws_comp.insert_cols(start_column_index + total_vol_idx)
    ws_comp.cell(row=start_row_index + 1, column=start_column_index + total_vol_idx).value = 'Total Volume (uL)\n(Count * Reaction Volume (uL) * 1.1 / 2)'
    # Insert 'Final Concentration (uM)' after 'Total Volume (uL)...'
    ws_comp.insert_cols(start_column_index + final_conc_idx)
    ws_comp.cell(row=start_row_index + 1, column=start_column_index + final_conc_idx).value = 'Final Concentration (uM)'
    # Insert 'Amount Needed (nmol)\n(Count * Reaction Volume (uL)\n* Final Concentration (uM) * 1.1 / 1000)' after Final Concentration
    ws_comp.insert_cols(start_column_index + amount_needed_idx)
    ws_comp.cell(row=start_row_index + 1, column=start_column_index + amount_needed_idx).value = 'Amount Needed (nmol)\n(Count * Reaction Volume (uL)\n* Final Concentration (uM) * 1.1 / 1000)'
    # Insert 'MW (Da)' as the last column
    ws_comp.insert_cols(start_column_index + mw_idx)
    ws_comp.cell(row=start_row_index + 1, column=start_column_index + mw_idx).value = 'MW (Da)'
    # Insert 'Mass Required (ug)' as the next column
    ws_comp.insert_cols(start_column_index + mass_required_idx)
    ws_comp.cell(row=start_row_index + 1, column=start_column_index + mass_required_idx).value = 'Mass Required (mg)\n(MW (Da) * Amount Needed (nmol) / 1000)'
    # Set values for all rows except header
    for row in range(start_row+1, ws_comp.max_row+1):
        comp = ws_comp.cell(row=row, column=start_column_index + 1).value
        # Set the 'Donor' and 'Donor Id' based on the component
        reagent = ws_comp.cell(row=row, column=start_column_index + donor_idx - 2).value
        reagent_id = prep_header_dict.get(reagent)
        cleaned_id = reagent_id.replace("Enzyme", "").replace("Donor: ", "").strip()
        ws_comp.cell(row=row, column=start_column_index + donor_idx).value = cleaned_id
        # Set the 'Reaction Volume (uL)' based on the default final concentrations
        reaction_volume = default_final_concentrations.get('Reaction Volume [uL]', 200)
        ws_comp.cell(row=row, column=start_column_index + reaction_vol_idx).value = reaction_volume
        # Set the 'Count' value
        count_cell = ws_comp.cell(row=row, column=start_column_index + count_idx).coordinate  # 'Count' now at col 2
        reaction_vol_cell = ws_comp.cell(row=row, column=start_column_index + reaction_vol_idx).coordinate  # 'Reaction Volume (uL)' now at col 3
        # Set Total Volume (uL) as Excel formula: (Count * Reaction Volume (uL) * 1.1) / 2
        ws_comp.cell(row=row, column=start_column_index + total_vol_idx).value = f"=({count_cell}*{reaction_vol_cell}*1.1)/2"
        final_conc_cell = ws_comp.cell(row=row, column=start_column_index + final_conc_idx).coordinate
        # Set Final Concentration (uM) based on the default final concentrations
        ws_comp.cell(row=row, column=start_column_index + amount_needed_idx).value = f"=({count_cell}*{reaction_vol_cell}*{final_conc_cell}*1.1/1000)"
        ws_comp.cell(row=row, column=start_column_index + final_conc_idx).value = default_final_concentrations.get(comp, '')
        # Set mass value in the last column
        ws_comp.cell(row=row, column=start_column_index + mw_idx).value = masses.get(comp, '')
        # Set Mass Required (ug) as Excel formula: (MW (Da) * Amount Needed (nmol)) / 1,000,000, rounded to 3 decimals
        mass_cell = ws_comp.cell(row=row, column=start_column_index + mw_idx).coordinate
        amount_needed_cell = ws_comp.cell(row=row, column=start_column_index + amount_needed_idx).coordinate
        ws_comp.cell(row=row, column=start_column_index + mass_required_idx).value = f"=IF(AND(ISNUMBER({mass_cell}),ISNUMBER({amount_needed_cell})),ROUND({mass_cell}*{amount_needed_cell}/1000000,3),"")"
    # Header formatting: bold, centered, bordered, wrap text
    from openpyxl.styles import Font, Border, Side, Alignment
    header_font = Font(bold=True)
    border = Border(bottom=Side(border_style="medium")) # if you want more borders insert here, left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"))
    for col in range(start_column_index + 1, ws_comp.max_column + 1):
        cell = ws_comp.cell(row=start_row_index + 1, column=col)
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    # Auto-fit all columns to the width of their titles or widest value
    # Auto-fit all columns to the width of their titles or widest value, accounting for wrapping
    for index, col in enumerate(ws_comp.columns):
        if index < start_column_index:
            continue # Skip the columns before the start_column_index
        header_cell = col[start_column_index]
        col_letter = header_cell.column_letter
        # Special case for 'Mass Required (mg)\n(MW (Da) * Amount Needed (nmol) / 1000)' column
        if header_cell.value == 'Mass Required (mg)\n(MW (Da) * Amount Needed (nmol) / 1000)':
            ws_comp.column_dimensions[col_letter].width = 30  # Set a tighter width for this column
            continue
        if '\n' in str(header_cell.value):
            lines = str(header_cell.value).split('\n')
            max_length = max(len(line) for line in lines)
        else:
            max_length = len(str(header_cell.value)) if header_cell.value is not None else 0
        for cell in col[start_column_index + 1:]:
            cell_length = len(str(cell.value)) if cell.value is not None else 0
            if cell_length > max_length:
                max_length = cell_length
        adjusted_width = max_length + 2
        ws_comp.column_dimensions[col_letter].width = adjusted_width




        
    # --- Create new worksheet for base reagent preparation in the same workbook ---
    # Remove existing 'Base Reagent Prep' sheet if it exists
    if 'Base Reagent Prep' in wb.sheetnames:
        std = wb['Base Reagent Prep']
        wb.remove(std)
    ws_prep = wb.create_sheet('Base Reagent Prep')
    # Place initial concentrations for hUba1, TCEP, ATP/Mg2+ in the top left corner with units, starting at cell B2
    start_row, start_col = 2, 2  # B2
    ws_prep.cell(row=start_row, column=start_col).value = 'Stock Conc.'
    ws_prep.cell(row=start_row+1, column=start_col).value = 'hUba1 (uM)'
    ws_prep.cell(row=start_row+2, column=start_col).value = 'TCEP (mM)'
    ws_prep.cell(row=start_row+3, column=start_col).value = 'ATP/Mg2+ (mM)'
    ws_prep.cell(row=start_row+1, column=start_col+1).value = initial_concentraions.get('hUba1', '') if 'hUba1' in initial_concentraions else ''
    ws_prep.cell(row=start_row+2, column=start_col+1).value = initial_concentraions.get('TCEP', '')
    ws_prep.cell(row=start_row+3, column=start_col+1).value = initial_concentraions.get('ATP/Mg2+', '')
    # Expand columns and center the writing
    from openpyxl.styles import Alignment, Border, Side, Font
    for row in range(start_row, start_row+4):
        for col in range(start_col, start_col+2):
            cell = ws_prep.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True) if row == start_row else Font(bold=False)
    # Set column widths for better appearance
    ws_prep.column_dimensions['B'].width = 18
    ws_prep.column_dimensions['C'].width = 14
    # Add a medium border under the title row, but remove the right border from the value column
    for col in range(start_col, start_col+2):
        cell = ws_prep.cell(row=start_row, column=col)
        if col == start_col+1:
            cell.border = Border(
                top=cell.border.top,
                left=cell.border.left,
                right=Side(border_style=None),
                bottom=Side(border_style="medium")
            )
        else:
            cell.border = Border(
                top=cell.border.top,
                left=cell.border.left,
                right=cell.border.right,
                bottom=Side(border_style="medium")
            )
    # Remove the right border from all cells in value column
    for row in range(start_row, start_row+4):
        cell = ws_prep.cell(row=row, column=start_col+1)
        cell.border = Border(
            top=cell.border.top,
            left=cell.border.left,
            right=Side(border_style=None),
            bottom=cell.border.bottom
        )
    # Draw a medium border around the whole table (B2:C5)
    for row in range(start_row, start_row+4):
        for col in range(start_col, start_col+2):
            cell = ws_prep.cell(row=row, column=col)
            border_args = {}
            if row == start_row:
                border_args['top'] = Side(border_style="medium")
            if row == start_row+3:
                border_args['bottom'] = Side(border_style="medium")
            if col == start_col:
                border_args['left'] = Side(border_style="medium")
            if col == start_col+1:
                border_args['right'] = Side(border_style="medium")
            # Merge with any existing border
            cell.border = Border(
                left=border_args.get('left', cell.border.left),
                right=border_args.get('right', cell.border.right),
                top=border_args.get('top', cell.border.top),
                bottom=border_args.get('bottom', cell.border.bottom)
            )

    # Add a new table under Stock Conc. in B7 for hUba1 needs
    hUba1_needs_start_row = 7  # Row 7 (B7)
    hUba1_needs_start_col = 2  # Column B
    hUba1_needs_labels = [
        'Total hUba1 Needs',
        'Volume (uL)',
        'nmol',
        'MW (Da)',
        'Mass (mg)'
    ]
    for i, label in enumerate(hUba1_needs_labels):
        ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col).value = label
        # Only bold the first label ('Total hUba1 Needs'), others not bold
        ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col).font = Font(bold=(i == 0))
        # Center the writing in column B7 to B11 (label column)
        ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col).alignment = Alignment(horizontal='center', vertical='center')
        # For 'Volume (uL)' row, insert formula summing G14, G32, G50
        if label == 'Volume (uL)':
            ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col + 1).value = '=G14+G32+G50'
        # For 'nmol' row, insert formula: =C3*C8/1000
        elif label == 'nmol':
            ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col + 1).value = '=C3*C8/1000'
        # For 'MW (Da)' row, insert value from masses['hUba1']
        elif label == 'MW (Da)':
            ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col + 1).value = masses['hUba1']
        # For 'Mass (mg)' row, insert formula: =ROUND(C9*C10/1E6, 3)
        elif label == 'Mass (mg)':
            ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col + 1).value = '=ROUND(C9*C10/1E6, 3)'
        else:
            ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col + 1).value = ''
        ws_prep.cell(row=hUba1_needs_start_row + i, column=hUba1_needs_start_col + 1).alignment = Alignment(horizontal='center', vertical='center')
    # Draw a border around the hUba1 needs box (B7:C10) and underline under the top row (B7:C7)
    from openpyxl.styles import Border, Side
    border_side = Side(border_style="medium")
    underline_side = Side(border_style="medium")
    for row in range(hUba1_needs_start_row, hUba1_needs_start_row + 5):
        for col in range(hUba1_needs_start_col, hUba1_needs_start_col + 2):
            cell = ws_prep.cell(row=row, column=col)
            border_args = {}
            if row == hUba1_needs_start_row:
                border_args['top'] = border_side
                border_args['bottom'] = underline_side  # underline under top row
            if row == hUba1_needs_start_row + 4:
                border_args['bottom'] = border_side
            if col == hUba1_needs_start_col:
                border_args['left'] = border_side
            if col == hUba1_needs_start_col + 1:
                border_args['right'] = border_side
            # Merge with any existing border
            cell.border = Border(
                left=border_args.get('left', cell.border.left),
                right=border_args.get('right', cell.border.right),
                top=border_args.get('top', cell.border.top),
                bottom=border_args.get('bottom', cell.border.bottom)
            )
    # --- Function to create the Base Reagent Prep table at a given location for a given reagent ---
    def create_base_reagent_prep_table(ws_prep, ws_comp, reagent, top_left_row, top_left_col):
        """
        Create the Base Reagent Prep table for a given reagent at the specified top-left cell (row, col).
        """


        base_reagent = prep_header_dict.get(reagent)

        prep_headers = [
            base_reagent,
            'Total Volume (uL)',
            'Amount Needed (nmol)'
        ]

        from openpyxl.styles import Border, Side, Font, Alignment
        right_border = Border(right=Side(border_style="thin"))
        # Write headers
        for i, header in enumerate(prep_headers):
            ws_prep.cell(row=top_left_row + i, column=top_left_col).value = header
            ws_prep.cell(row=top_left_row + i, column=top_left_col).font = Font(bold=True)
            ws_prep.cell(row=top_left_row + i, column=top_left_col).border = right_border
            ws_prep.cell(row=top_left_row + i, column=top_left_col).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        
        # Add header for Volume (uL) above the third column
        ws_prep.cell(row=top_left_row, column=top_left_col+2).value = 'Volume (uL)'
        ws_prep.cell(row=top_left_row, column=top_left_col+2).font = Font(bold=True)
        ws_prep.cell(row=top_left_row, column=top_left_col+2).alignment = Alignment(horizontal='center', vertical='center')
        
        # Remove any old header above (only if row > 1)
        if top_left_row > 1:
            ws_prep.cell(row=top_left_row-1, column=top_left_col+2).value = None

        # Write values (formulas) for the reagent
        found = False
        for comp_row in range(2, ws_comp.max_row+1):
            # TODO Here is one of two places where we write the an individual column number
            # In this case it is column 2, which is the second column in the ws_comp sheet
            # This is prone to errors if the column order changes
            comp_name = ws_comp.cell(row=comp_row, column=2).value
            if comp_name == reagent:
                # Here is the second place where we write the component name - B, E, G - prone to errors
                # Write the component name in the first column of the prep table
                ws_prep.cell(row=top_left_row, column=top_left_col+1).value = f"='Base Reagent Counts'!B{comp_row}"
                ws_prep.cell(row=top_left_row, column=top_left_col+1).alignment = Alignment(horizontal='center', vertical='center')
                ws_prep.cell(row=top_left_row, column=top_left_col+1).font = Font(bold=True)
                ws_prep.cell(row=top_left_row+1, column=top_left_col+1).value = f"='Base Reagent Counts'!F{comp_row}"
                ws_prep.cell(row=top_left_row+2, column=top_left_col+1).value = f"='Base Reagent Counts'!H{comp_row}"
                found = True
                break
        if not found:
            ws_prep.cell(row=top_left_row, column=top_left_col+1).value = reagent
            ws_prep.cell(row=top_left_row, column=top_left_col+1).alignment = Alignment(horizontal='center', vertical='center')
            ws_prep.cell(row=top_left_row, column=top_left_col+1).font = Font(bold=True)
            ws_prep.cell(row=top_left_row+1, column=top_left_col+1).value = ''
            ws_prep.cell(row=top_left_row+2, column=top_left_col+1).value = ''
        # Add 6 new rows for stock conc. and stock volume for 1, 2, 3, with default value 0
        extra_labels = [
            'Stock Conc. 1 (uM)', 'Stock Volume 1 (uL)',
            'Stock Conc. 2 (uM)', 'Stock Volume 2 (uL)',
            'Stock Conc. 3 (uM)', 'Stock Volume 3 (uL)'
        ]
        for i, label in enumerate(extra_labels):
            row = top_left_row + 3 + i
            ws_prep.cell(row=row, column=top_left_col).value = label
            ws_prep.cell(row=row, column=top_left_col).font = Font(bold=True)
            ws_prep.cell(row=row, column=top_left_col).border = right_border
            ws_prep.cell(row=row, column=top_left_col).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            ws_prep.cell(row=row, column=top_left_col+1).value = 0
        # Write formulas for stock volumes in the volume column (col 3)
        ws_prep.cell(row=top_left_row+4, column=top_left_col+2).value = f"=IF(AND({chr(65+top_left_col)}{top_left_row+3}>0,{chr(65+top_left_col)}{top_left_row+2}>0),MIN({chr(65+top_left_col)}{top_left_row+4},ROUND({chr(65+top_left_col)}{top_left_row+2}*1000/{chr(65+top_left_col)}{top_left_row+3},2)),0)"
        ws_prep.cell(row=top_left_row+6, column=top_left_col+2).value = f"=IF(AND({chr(65+top_left_col)}{top_left_row+5}>0,MAX(0,{chr(65+top_left_col)}{top_left_row+2}-{chr(65+top_left_col+1)}{top_left_row+4}*{chr(65+top_left_col)}{top_left_row+3}/1000)>0),MIN({chr(65+top_left_col)}{top_left_row+6},ROUND(MAX(0,{chr(65+top_left_col)}{top_left_row+2}-{chr(65+top_left_col+1)}{top_left_row+4}*{chr(65+top_left_col)}{top_left_row+3}/1000)*1000/{chr(65+top_left_col)}{top_left_row+5},2)),0)"
        ws_prep.cell(row=top_left_row+8, column=top_left_col+2).value = f"=IF(AND({chr(65+top_left_col)}{top_left_row+7}>0,MAX(0,{chr(65+top_left_col)}{top_left_row+2}-{chr(65+top_left_col+1)}{top_left_row+4}*{chr(65+top_left_col)}{top_left_row+3}/1000-{chr(65+top_left_col+1)}{top_left_row+6}*{chr(65+top_left_col)}{top_left_row+5}/1000)>0),MIN({chr(65+top_left_col+2)}{top_left_row+8},ROUND(MAX(0,{chr(65+top_left_col)}{top_left_row+2}-{chr(65+top_left_col+1)}{top_left_row+4}*{chr(65+top_left_col)}{top_left_row+3}/1000-{chr(65+top_left_col+1)}{top_left_row+6}*{chr(65+top_left_col)}{top_left_row+5}/1000)*1000/{chr(65+top_left_col)}{top_left_row+7},2)),0)"
        # At the bottom, add a cell for 'Required nmol needed' and formula
        ws_prep.cell(row=top_left_row+9, column=top_left_col).value = 'Required nmol needed'
        ws_prep.cell(row=top_left_row+9, column=top_left_col).font = Font(bold=True)
        ws_prep.cell(row=top_left_row+9, column=top_left_col).border = right_border
        ws_prep.cell(row=top_left_row+9, column=top_left_col).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        ws_prep.cell(row=top_left_row+9, column=top_left_col+1).value = f"=ROUND(MAX(0,{chr(65+top_left_col)}{top_left_row+2}-{chr(65+top_left_col+1)}{top_left_row+4}*{chr(65+top_left_col)}{top_left_row+3}/1000-{chr(65+top_left_col+1)}{top_left_row+6}*{chr(65+top_left_col)}{top_left_row+5}/1000-{chr(65+top_left_col+1)}{top_left_row+8}*{chr(65+top_left_col)}{top_left_row+7}/1000),2)"
        # Add a border between the second and third column for all relevant rows
        mid_right_border = Border(right=Side(border_style="thin"))
        for i in range(0, 10):
            ws_prep.cell(row=top_left_row + i, column=top_left_col+1).border = mid_right_border
        # Auto-fit columns for this table: set all columns to the width of the first column
        first_col_letter = ws_prep.cell(row=top_left_row, column=top_left_col).column_letter
        first_col_width = max(len(str(ws_prep.cell(row=top_left_row + i, column=top_left_col).value)) for i in range(0, 10)) + 2
        for col in range(top_left_col, top_left_col+3):
            col_letter = ws_prep.cell(row=top_left_row, column=col).column_letter
            ws_prep.column_dimensions[col_letter].width = first_col_width
        # Place the Additional Reagent section below the main table, with only one blank row for spacing
        add_reagent_start_row = top_left_row + 10  # 10 rows for main table
        ws_prep.cell(row=add_reagent_start_row, column=top_left_col).value = None
        # Now add the additional reagents table below the single blank row
        add_table_row = add_reagent_start_row + 1
        if reagent in ['Ubc13/Mms2', 'gp78/Ube2g2', 'Ube2K']:
            extra_reagents = ['hUba1 (uM)', 'TCEP (mM)', 'ATP/Mg2+ (mM)']
        else:
            extra_reagents = ['TCEP (mM)', 'ATP/Mg2+ (mM)']
        extra_col = top_left_col
        ws_prep.cell(row=add_table_row, column=extra_col).value = 'Reagent'
        ws_prep.cell(row=add_table_row, column=extra_col+1).value = 'Final Conc.'
        ws_prep.cell(row=add_table_row, column=extra_col+2).value = 'Volume (uL)'
        # Style header row for second table
        for j, label in enumerate(['Reagent', 'Final Conc.', 'Volume (uL)']):
            cell = ws_prep.cell(row=add_table_row, column=extra_col+j)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Fill in second table and style cells
        for i, reagent_name in enumerate(extra_reagents):
            for j in range(3):
                cell = ws_prep.cell(row=add_table_row+1+i, column=extra_col+j)
                # Only the first column (Reagent) is bold, others are not
                if j == 0:
                    cell.font = Font(bold=True)
                else:
                    cell.font = Font(bold=False)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_prep.cell(row=add_table_row+1+i, column=extra_col).value = reagent_name
            # Final Conc.
            if reagent_name == 'hUba1 (uM)':
                ws_prep.cell(row=add_table_row+1+i, column=extra_col+1).value = default_final_concentrations.get('hUba1', '')
            elif reagent_name == 'TCEP (mM)':
                ws_prep.cell(row=add_table_row+1+i, column=extra_col+1).value = default_final_concentrations.get('TCEP', '')
            elif reagent_name == 'ATP/Mg2+ (mM)':
                ws_prep.cell(row=add_table_row+1+i, column=extra_col+1).value = default_final_concentrations.get('ATP/Mg2+', '')
            # Set Volume (uL) formula
            comp_vol_cell = ws_prep.cell(row=top_left_row+1, column=top_left_col+1).coordinate
            if reagent_name == 'hUba1 (uM)':
                init_conc_cell = ws_prep.cell(row=3, column=3).coordinate  # C3
            elif reagent_name == 'TCEP (mM)':
                init_conc_cell = ws_prep.cell(row=4, column=3).coordinate  # C4
            elif reagent_name == 'ATP/Mg2+ (mM)':
                init_conc_cell = ws_prep.cell(row=5, column=3).coordinate  # C5
            else:
                init_conc_cell = None
            final_conc_cell = ws_prep.cell(row=add_table_row+1+i, column=extra_col+1).coordinate
            if init_conc_cell:
                if reagent_name == 'hUba1 (uM)':
                    ws_prep.cell(row=add_table_row+1+i, column=extra_col+2).value = f"=IF({init_conc_cell}>0,({final_conc_cell}*2*{comp_vol_cell})/{init_conc_cell},0)"
                else:
                    ws_prep.cell(row=add_table_row+1+i, column=extra_col+2).value = f"=IF({init_conc_cell}>0,({final_conc_cell}*{comp_vol_cell})/{init_conc_cell},0)"
            else:
                ws_prep.cell(row=add_table_row+1+i, column=extra_col+2).value = ''
        # Add HEPES buffer row
        hepes_row = add_table_row+1+len(extra_reagents)
        for j in range(3):
            cell = ws_prep.cell(row=hepes_row, column=extra_col+j)
            # Only the first column (Reagent) is bold, others are not
            if j == 0:
                cell.font = Font(bold=True)
            else:
                cell.font = Font(bold=False)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_prep.cell(row=hepes_row, column=extra_col).value = 'HEPES buffer'
        ws_prep.cell(row=hepes_row, column=extra_col+1).value = ''
        # Build formula for HEPES buffer volume
        # Base Reagent volume cell
        comp_vol_cell = ws_prep.cell(row=top_left_row+1, column=top_left_col+1).coordinate
        # Collect all other volume cells in both tables
        volume_cells = []
        # Main table: stock volumes (every other row, starting from top_left_row+4, col=top_left_col+2, 3 stock volumes)
        for k in [4,6,8]:
            volume_cells.append(ws_prep.cell(row=top_left_row+k, column=top_left_col+2).coordinate)
        # Additional reagents table: hUba1, TCEP, ATP/Mg2+ (rows add_table_row+1 to add_table_row+len(extra_reagents), col=extra_col+2)
        for k in range(len(extra_reagents)):
            volume_cells.append(ws_prep.cell(row=add_table_row+1+k, column=extra_col+2).coordinate)
        sum_formula = '+'.join(volume_cells)
        ws_prep.cell(row=hepes_row, column=extra_col+2).value = f"={comp_vol_cell}-({sum_formula})"
        # --- Draw a single, unbroken medium border around both tables as one block ---
        from openpyxl.styles import Border, Side
        outer_side = Side(border_style="medium")  # less bold than thick
        inner_side = Side(border_style="thin")

        # Add a light bold (medium) border under the title line of the component calcualtion table & additional reagents table
        for row_num in [top_left_row, add_table_row, add_table_row -1]:
            for j in range(3):
                cell = ws_prep.cell(row=row_num, column=extra_col + j)
                cell.border = Border(
                    top=cell.border.top,
                    left=cell.border.left,
                    right=cell.border.right,
                    bottom=Side(border_style="medium")
                )

        # Rectangle: from top_left_row, top_left_col to bottom row/col of second table
        # Bottom border should be directly under HEPES buffer (not after the extra blank row)
        total_rows = (10 + 1 + 1 + len(extra_reagents) + 1)  # main table + blank + header + extra reagents + HEPES
        start_row = top_left_row
        end_row = hepes_row  # last row with content is HEPES buffer
        start_col = top_left_col
        end_col = top_left_col + 2
        # Add a blank row after HEPES buffer for visual separation
        for c in range(start_col, end_col+1):
            ws_prep.cell(row=hepes_row+1, column=c).value = None
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws_prep.cell(row=r, column=c)
                border_args = {}
                # Outer border
                if r == start_row:
                    border_args['top'] = outer_side
                if r == end_row:
                    border_args['bottom'] = outer_side
                if c == start_col:
                    border_args['left'] = outer_side
                if c == end_col:
                    border_args['right'] = outer_side
                # For the bottom (additional reagents) table, add thin vertical lines between columns (now up to HEPES buffer)
                if r >= add_table_row and r <= end_row:
                    if c == start_col + 1:
                        border_args['left'] = inner_side
                    if c == start_col + 1:
                        border_args['right'] = inner_side
                if border_args:
                    existing = cell.border
                    cell.border = Border(
                        left=border_args.get('left', existing.left),
                        right=border_args.get('right', existing.right),
                        top=border_args.get('top', existing.top),
                        bottom=border_args.get('bottom', existing.bottom)
                    )
    # Create the Base Reagent Prep table for each reagent at specified locations
    # Only create the Base Reagent Prep table for 'Ubc13/Mms2' at (5, 4)
    create_base_reagent_prep_table(ws_prep, ws_comp, 'Ubc13/Mms2', 2, 5)
    create_base_reagent_prep_table(ws_prep, ws_comp, 'gp78/Ube2g2', 20, 5)
    create_base_reagent_prep_table(ws_prep, ws_comp, 'Ube2K', 38, 5)

    # Create the Base Reagent Prep table for the donors
    create_base_reagent_prep_table(ws_prep, ws_comp, 'ubi_ubq_1_K48_ABOC_K63_ABOC', 74, 10) 
    create_base_reagent_prep_table(ws_prep, ws_comp, 'ubi_ubq_1_K48_SMAC_K63_ABOC', 20, 10)  
    create_base_reagent_prep_table(ws_prep, ws_comp, 'ubi_ubq_1_K48_ABOC_K63_SMAC', 56, 10) 
    create_base_reagent_prep_table(ws_prep, ws_comp, 'ubi_ubq_1_K48_SMAC', 2, 10) 
    create_base_reagent_prep_table(ws_prep, ws_comp, 'ubi_ubq_1_K63_SMAC', 38, 10) 
    
    # Widen the 11th column (column 'K') of the 'Base Reagent Prep' worksheet for better readability
    ws_prep = wb['Base Reagent Prep']
    ws_prep.column_dimensions['K'].width = 28

    # --- New: Add Acceptor Counts worksheet if acceptors_96_counts is provided ---
    def add_acceptor_counts_sheet(wb, acceptors_96_counts):
        """
        Add a worksheet displaying acceptor counts (acceptors_96_counts dict) to the workbook.
        Also display dimer_encoded_dictionary as a table using pandas, starting at D2.
        """

        # Check if acceptors_96_counts is provided and not empty
        ws_acceptor = wb.create_sheet('Acceptor Counts')
        
        column_start = 2  # Start at column B

        # --- Display merged dimer_encoded_dictionary and acceptors_96_counts as a table using pandas, starting at D2 ---
        if 'dimer_encoded_dictionary' in globals() and dimer_encoded_dictionary and acceptors_96_counts:
            import pandas as pd
            # dimer_encoded_dictionary: {encoded_value: dimer_key}
            # acceptors_96_counts: {encoded_value: count}
            merged = []
            for dimer_name, encoded_value in dimer_encoded_dictionary.items():
                count = acceptors_96_counts.get(encoded_value, 0)
                merged.append({'Encoded Value': encoded_value, 'Acceptor': dimer_name, 'Count': count})

            # Merge acceptor_description_dict into merged based on Encoded Value <-> No.
            # Build a lookup from number to description dict
            desc_by_number = {}
            for desc in acceptor_description_dict:
                try:
                    num = int(desc["No."].split()[-1])
                    desc_by_number[num] = desc
                except Exception:
                    continue
            for row in merged:
                encoded_val = row['Encoded Value']
                desc = desc_by_number.get(encoded_val)
                if desc:
                    for k, v in desc.items():
                        if k != "No." and k not in row:
                            row[k] = v

            merged_df = pd.DataFrame(merged)
            # Add 'Acceptor No.' column from acceptor_description_dict if possible
            acceptor_no_map = {int(desc["No."].split()[-1]): desc["No."] for desc in acceptor_description_dict}
            merged_df.insert(1, 'Acceptor No.', merged_df['Encoded Value'].map(acceptor_no_map))
            # Remove the 'Acceptor' column if present
            if 'Acceptor' in merged_df.columns:
                merged_df = merged_df.drop(columns='Acceptor')
            # Move 'Count' column to the end
            if 'Count' in merged_df.columns:
                count_col = merged_df.pop('Count')
                merged_df['Count'] = count_col
            # Add 'nmol / well' and 'Total nmol' columns after 'Count'
            nmol_per_well_default = 10
            if 'Count' in merged_df.columns:
                count_idx = merged_df.columns.get_loc('Count')
                merged_df.insert(count_idx + 1, 'nmol / well', nmol_per_well_default)
                merged_df.insert(count_idx + 2, 'Total nmol', '')  # Placeholder for Excel formula
            # Add 'Stock Conc. (uM)', 'Volume / well', and 'Total Volume needed' columns at the end
            stock_conc_default = 20
            merged_df['Stock Conc. (uM)'] = stock_conc_default
            merged_df['Volume / well'] = ''  # Placeholder for Excel formula
            merged_df['Total Volume needed'] = ''  # Placeholder for Excel formula

            # Write headers (including new columns)
            for col_idx, col_name in enumerate(merged_df.columns, start=column_start):
                ws_acceptor.cell(row=2, column=col_idx).value = col_name
                ws_acceptor.cell(row=2, column=col_idx).font = header_font
                ws_acceptor.cell(row=2, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
            # Add a bold line (medium border) under the title line
            from openpyxl.styles import Border, Side
            for col_idx in range(column_start, column_start + len(merged_df.columns)):
                cell = ws_acceptor.cell(row=2, column=col_idx)
                cell.border = Border(
                    top=cell.border.top,
                    left=cell.border.left,
                    right=cell.border.right,
                    bottom=Side(border_style="medium")
                )
            # Write data and set Excel formulas for new columns
            for row_idx, row in enumerate(merged_df.itertuples(index=False), start=3):
                for col_idx, value in enumerate(row, start=column_start):
                    col_name = merged_df.columns[col_idx - column_start]
                    if col_name == 'Total nmol':
                        count_col_letter = ws_acceptor.cell(row=2, column=col_idx-2).column_letter
                        nmol_col_letter = ws_acceptor.cell(row=2, column=col_idx-1).column_letter
                        ws_acceptor.cell(row=row_idx, column=col_idx).value = f"={count_col_letter}{row_idx}*{nmol_col_letter}{row_idx}"
                    elif col_name == 'Volume / well':
                        nmol_col_letter = ws_acceptor.cell(row=2, column=col_idx-3).column_letter
                        stock_col_letter = ws_acceptor.cell(row=2, column=col_idx-1).column_letter
                        ws_acceptor.cell(row=row_idx, column=col_idx).value = f"=ROUND({nmol_col_letter}{row_idx}/{stock_col_letter}{row_idx}*1000,2)"
                    elif col_name == 'Total Volume needed':
                        total_nmol_col_letter = ws_acceptor.cell(row=2, column=col_idx-3).column_letter
                        stock_col_letter = ws_acceptor.cell(row=2, column=col_idx-2).column_letter
                        ws_acceptor.cell(row=row_idx, column=col_idx).value = f"=ROUND({total_nmol_col_letter}{row_idx}/{stock_col_letter}{row_idx}*1000,2)"
                    else:
                        ws_acceptor.cell(row=row_idx, column=col_idx).value = value
                    ws_acceptor.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
            # Auto-fit all columns to the width of their header text
            for col_idx, col_name in enumerate(merged_df.columns, start=column_start):
                col_letter = ws_acceptor.cell(row=2, column=col_idx).column_letter
                header_length = len(str(col_name))
                ws_acceptor.column_dimensions[col_letter].width = header_length + 2

    # At the end of create_combined_xlsx, add acceptor counts sheet if provided
    if 'acceptors_96_counts' in globals() and acceptors_96_counts:
        add_acceptor_counts_sheet(wb, acceptors_96_counts)

    # When done, save back to BytesIO for FastAPI
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    return final_output

excel_bytes = create_combined_xlsx_bytes(mixture_counts, component_counts, 'mixture_and_component_counts.xlsx', all_components=components, e_d_encoded_dictionary=e_d_encoded_dictionary)


# 

# Suppose 'excel_bytes' is your BytesIO object
with open("mixture_and_component_counts.xlsx", "wb") as f:
    f.write(excel_bytes.getbuffer())

# ============================================================
# Close Excel file and open it
# ============================================================

def close_excel_file(filename):
    """
    Ensure the Excel file is closed before proceeding.
    """
    import os
    import time
    # Only attempt on macOS
    if sys.platform != 'darwin':
        return
    # Check if Excel is running
    def is_excel_running():
        try:
            result = subprocess.run([
                'osascript',
                '-e', 'tell application "System Events" to (name of processes) contains "Microsoft Excel"'
            ], capture_output=True, text=True)
            return 'true' in result.stdout.lower()
        except Exception:
            return False
    # Try to close the file if already open (macOS only, using AppleScript)
    if is_excel_running():
        closed = False
        for _ in range(10):  # Try up to 10 times
            try:
                subprocess.run([
                    'osascript',
                    '-e', f'tell application "Microsoft Excel" to close (every document whose name is "{filename}")',
                    '-e', 'delay 0.5'
                ], check=True)
                closed = True
                break
            except Exception:
                time.sleep(0.5)
        if not closed:
            print(f"Warning: Could not close Excel file {filename} automatically.")
        # Wait a moment to ensure file is closed
        time.sleep(1.5)

def close_excel_completely():
    """
    Ensure Microsoft Excel is fully quit before opening the file again.
    """
    import time
    import subprocess
    if sys.platform != 'darwin':
        return
    try:
        subprocess.run([
            'osascript',
            '-e', 'tell application "Microsoft Excel" to quit saving no'
        ], check=True)
        time.sleep(2)  # Wait to ensure Excel has fully quit
    except Exception:
        print("Warning: Could not quit Microsoft Excel automatically.")

close_excel_file('mixture_and_component_counts.xlsx')
close_excel_completely()
subprocess.run(['open', 'mixture_and_component_counts.xlsx'])
